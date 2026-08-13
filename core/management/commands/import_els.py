"""
downloads 폴더의 청약중인상품_*.xlsx를 감지해 Product로 임포트.

- ALL 시트만 파싱 (다른 시트는 부분집합)
- 열은 헤더 이름으로 찾는다 (아래 COLUMNS 주석 참고)
- ImportLog로 동일 파일 재처리 방지
- 프리셋 매칭 신규 상품 텔레그램 알림 (NotifiedMatch로 중복 방지)
- 보유 Investment 평가일 D-1 알림 (RedemptionAlert로 중복 방지)
- 월~수 새 파일 없으면 목요일에 리마인더
"""

import datetime
import glob
import os
import re
from datetime import date, timedelta

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand

from core import notify, parsers, telegram
from core.models import ImportLog, Product

# ELS_Curator 엑셀은 주차마다 열 구성이 다르다. 실제로 downloads에 쌓인 42개 파일에서
# 서로 다른 레이아웃이 15종 나왔다 — 열이 12~21개로 오가고, 같은 자리(11번)에
# 상품유형·구분·ki·KI·낙인조건·micron여부가 번갈아 들어온다.
# 열 번호를 고정하면 엉뚱한 값을 읽는다. 실제로 2026-07-16 시딩 배치에서
# 청약종료일 열이 아예 없는 파일(20260616_0925)을 인덱스로 읽어 sub_end=NULL 369건이
# 들어갔고, 그게 (issuer, product_no, sub_end) 유니크 제약을 빠져나가
# 중복 274쌍을 만들었다. 그래서 헤더 이름으로 찾고, 필수 열이 없으면 파일을 버린다.
# (2026-08-05)
COLUMNS = {
    "issuer": ("발행회사",),
    "name": ("상품명",),
    "assets": ("기초자산",),
    "issue_date": ("발행일",),
    "expiry_date": ("만기일",),
    "yield_rate": ("연수익률",),
    "max_loss": ("최대손실률",),          # 없는 파일이 있다 (선택)
    "sub_start": ("청약시작일",),          # 없는 파일이 있다 (선택)
    "sub_end": ("청약종료일", "청약마감일"),
    "desc": ("상품유형",),
}
# 이 열이 하나라도 없으면 임포트하지 않는다 — 없는 채로 넣으면 조용히 결손 데이터가 쌓인다.
# sub_start·max_loss는 키도 아니고 모델에서 null 허용이라 빠져도 그냥 비워 둔다.
REQUIRED_COLUMNS = (
    "issuer", "name", "assets", "issue_date", "expiry_date",
    "yield_rate", "sub_end", "desc",
)


def _norm_header(h):
    """헤더 비교용 정규화 — 공백·괄호·% 제거. '최대손실률(%)' → '최대손실률'."""
    return re.sub(r"[\s()%]", "", str(h or ""))


def _map_columns(header_row):
    """헤더 행 → {필드명: 열 인덱스}. 못 찾은 필드는 키가 없다."""
    found = {}
    for idx, cell in enumerate(header_row):
        norm = _norm_header(cell)
        if not norm:
            continue
        for field, aliases in COLUMNS.items():
            if field not in found and norm in aliases:
                found[field] = idx
    return found


def _to_date(val):
    """엑셀의 20260323 형식 int/str → date. 날짜 서식 셀(datetime)도 받는다."""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip().replace(".", "").replace("-", "")
    if len(s) == 8 and s.isdigit():
        try:
            return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
        except ValueError:
            return None
    return None


def _to_float(val):
    try:
        return float(str(val).replace(",", "").replace("%", "").strip())
    except (ValueError, TypeError):
        return None


class Command(BaseCommand):
    help = "ELS 엑셀 임포트 + 프리셋 매칭 알림 + 상환 평가일 알림"

    def add_arguments(self, parser):
        parser.add_argument("--file", help="특정 파일만 임포트 (경로)")
        parser.add_argument("--no-notify", action="store_true", help="텔레그램 발송 생략")

    def handle(self, *args, **opts):
        should_notify = not opts["no_notify"]

        if opts.get("file"):
            files = [opts["file"]]
        else:
            pattern = os.path.join(settings.ELS_DOWNLOADS_DIR, "청약중인상품_*.xlsx")
            # _수정/_서식테스트 등 가공본은 제외 (원본만 임포트)
            files = sorted(
                f for f in glob.glob(pattern)
                if "_수정" not in os.path.basename(f)
                and "테스트" not in os.path.basename(f)
            )

        processed = ImportLog.objects.values_list("filename", flat=True)
        new_files = [f for f in files if os.path.basename(f) not in processed]

        total_new_products = 0
        for path in new_files:
            n_rows, n_new = self._import_file(path)
            total_new_products += n_new
            self.stdout.write(f"[임포트] {os.path.basename(path)}: {n_rows}행 중 신규 {n_new}건")

        if not new_files:
            self.stdout.write("새 파일 없음")
            self._maybe_remind(should_notify)

        # 프리셋 매칭 알림
        if should_notify and total_new_products:
            notify.notify_preset_matches(self.stdout)

        # 상환 평가일 D-1 알림은 send_redemption_alert(아침 크론)로 분리

        if new_files and should_notify:
            telegram.send_message(
                f"[ELS 레이더] 임포트 완료\n"
                f"파일 {len(new_files)}개 / 신규 상품 {total_new_products}건\n"
                f"대시보드: {settings.SITE_URL}"
            )

    # ── 파일 임포트 ─────────────────────────────
    def _import_file(self, path):
        wb = openpyxl.load_workbook(path, data_only=True)
        if "ALL" not in wb.sheetnames:
            self.stderr.write(f"ALL 시트 없음: {path}")
            return 0, 0
        ws = wb["ALL"]

        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        col = _map_columns(header)
        missing = [f for f in REQUIRED_COLUMNS if f not in col]
        if missing:
            # 열이 빠진 파일을 인덱스로 밀어넣으면 결손·오염 데이터가 조용히 쌓인다.
            # ImportLog는 남겨 매 실행마다 같은 경보가 반복되는 것만 막는다.
            names = ", ".join(COLUMNS[f][0] for f in missing)
            self.stderr.write(f"[열 누락] {os.path.basename(path)}: {names} / 임포트 건너뜀")
            telegram.send_message(
                f"[임포트 중단] {os.path.basename(path)}\n"
                f"필수 열 누락: {names}\n"
                f"실제 헤더: {', '.join(str(h) for h in header if h)}\n"
                "엑셀을 다시 내려받아 주세요 (이 파일은 임포트하지 않았습니다)."
            )
            ImportLog.objects.create(
                filename=os.path.basename(path), row_count=0, new_count=0
            )
            return 0, 0

        def cell(row, field):
            i = col.get(field)
            return row[i] if i is not None and i < len(row) else None

        n_rows = n_new = n_skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            issuer = str(cell(row, "issuer") or "").strip()
            if not issuer:
                continue
            # 기초자산이 비어 유형(종목형/지수형) 분류가 불가한 결손 행은 버림
            # (미분류 상품이 목록·포트폴리오 유형 집계를 깨뜨리는 것 원천 차단)
            assets = str(cell(row, "assets") or "")
            if not parsers.classify_asset(assets):
                continue

            # sub_end는 유니크 키다. NULL로 넣으면 SQLite가 NULL끼리 다르다고 봐서
            # 제약을 통과하고 같은 상품이 여러 행으로 갈라진다 — 행을 버린다.
            sub_end = _to_date(cell(row, "sub_end"))
            if sub_end is None:
                n_skipped += 1
                continue
            n_rows += 1

            desc = str(cell(row, "desc") or "")
            issue_date = _to_date(cell(row, "issue_date"))
            expiry_date = _to_date(cell(row, "expiry_date"))
            ki = parsers.extract_ki(desc)
            barriers = parsers.extract_barriers(desc)
            period = parsers.extract_period(desc, issue_date, expiry_date, barriers)
            asset_type = parsers.classify_asset(assets) or ""

            is_no_ki = ki == "NoKI"
            ki_val = None if (ki is None or is_no_ki) else int(ki)

            # 엑셀 수입분은 상품명이 회차번호뿐이라 설명이 사실상 유일한 단서다.
            # 그래도 판정은 scrape_kofia·reparse_products와 같은 parsers 함수를 쓴다
            # (예전엔 이 파일에도 같은 분기가 복사돼 있어 셋이 따로 놀았다).
            product_type = parsers.classify_product_type(
                str(cell(row, "name") or "").strip(), desc, ki_val)

            # 발행통화도 같은 parsers 함수를 쓴다. 못 읽으면 None —
            # scrape_kofia와 같은 이유로 KRW로 단정하지 않는다(아래 defaults 참조).
            currency = parsers.currency_from_description(desc)

            defaults = dict(
                name=str(cell(row, "name") or "").strip(),
                product_type=product_type,
                yield_rate=_to_float(cell(row, "yield_rate")),
                max_loss=_to_float(cell(row, "max_loss")),
                ki=ki_val,
                is_no_ki=is_no_ki,
                barrier_first=int(barriers[0]) if barriers else None,
                barrier_last=int(barriers[-1]) if barriers else None,
                barriers_raw=[int(b) for b in barriers] if barriers else None,
                period_months=period,
                asset_type=asset_type,
                assets_raw=assets.strip(),
                issue_date=issue_date,
                expiry_date=expiry_date,
                sub_start=_to_date(cell(row, "sub_start")),
                description=desc,
            )
            # 설명이 통화를 밝힌 경우에만 쓴다. 안 밝히면 신규는 모델 기본값 KRW로
            # 만들어지고, 기존 행은 이미 확정된 값을 그대로 둔다.
            if currency:
                defaults["currency"] = currency

            _, created = Product.objects.update_or_create(
                issuer=issuer,
                product_no=str(cell(row, "name") or "").strip(),
                sub_end=sub_end,
                defaults=defaults,
            )
            if created:
                n_new += 1

        if n_skipped:
            self.stderr.write(
                f"[청약종료일 없음] {os.path.basename(path)}: {n_skipped}행 건너뜀"
            )

        ImportLog.objects.create(
            filename=os.path.basename(path), row_count=n_rows, new_count=n_new
        )
        return n_rows, n_new

    # ── 목요일 리마인더 ──────────────────────────
    def _maybe_remind(self, should_notify):
        today = date.today()
        if today.weekday() != 3:  # 목요일
            return
        monday = today - timedelta(days=3)
        recent = ImportLog.objects.filter(imported_at__date__gte=monday).exists()
        if not recent and should_notify:
            telegram.send_message(
                "[리마인더] 이번 주 ELS 데이터가 아직 없습니다.\n"
                "ELS_Curator를 실행하거나 자동수집(scrape_kofia)을 확인해주세요."
            )
