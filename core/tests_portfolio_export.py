"""보유계약 엑셀 다운로드(17열 양식) 테스트.

핵심은 두 가지다.
  ① 조 팀장이 실제로 쓰던 시트 6행을 운영 DB 값으로 그대로 재현하는가
  ② 남의 투자 기록이 절대 섞이지 않는가
①의 고정값은 2026-08-06 운영 DB(읽기 전용)에서 그대로 가져온 것이라,
매핑 규칙이 바뀌면 이 테스트가 먼저 깨진다.
"""

import io
from datetime import date

import openpyxl
from django.contrib.auth import get_user_model
from django.test import TestCase

from core import portfolio_export as pfx
from core.models import Investment, Product


def _product(**kw):
    kw.setdefault("product_type", "ELS")
    return Product.objects.create(**kw)


class ManagerSheetReproductionTest(TestCase):
    """조 팀장 실제 보유 6건을 한 칸씩 재현한다 (운영 DB 2026-08-06 값)."""

    def setUp(self):
        self.user = get_user_model().objects.create_user(username="jo", password="x")
        self.rows = [
            dict(issuer="삼성증권", product_no="30868", asset_type="지수형",
                 assets_raw="KOSPI200 , S&P500 , Nikkei225", yield_rate=20.0, ki=45,
                 barrier_first=90, barrier_last=75, period_months=3,
                 barriers_raw=[90, 90, 90, 90, 90, 90, 85, 85, 85, 80, 80, 75],
                 issue_date=date(2026, 4, 3), expiry_date=date(2029, 4, 2),
                 description="[스텝다운] 3년/3개월,45KI(90,90,90,90,90,90,85,85,85,80,80,75)%,세전 연 20%",
                 eval_dates=["2026-07-01", "2026-10-02", "2026-12-30", "2027-04-02",
                             "2027-07-02", "2027-10-01", "2027-12-30", "2028-03-31",
                             "2028-06-30", "2028-09-29", "2028-12-28", "2029-04-05"],
                 amount=30_000_000, status="조기상환"),
            dict(issuer="키움증권", product_no="1806", asset_type="종목형",
                 assets_raw="Palantir , Micron", yield_rate=24.2, ki=20,
                 barrier_first=80, barrier_last=65, period_months=6,
                 barriers_raw=[80, 80, 75, 75, 70, 65],
                 issue_date=date(2026, 4, 3), expiry_date=date(2029, 4, 5),
                 description="[스텝다운] 3년/6개월/80-80-75-75-70-65/KI20",
                 eval_dates=["2026-10-02", "2027-04-02", "2027-10-01", "2028-03-31",
                             "2028-10-02", "2029-04-05"],
                 amount=30_000_000, status="보유중"),
            dict(issuer="삼성증권", product_no="30895", asset_type="종목형",
                 assets_raw="현대차, 삼성전자", yield_rate=19.8, ki=25,
                 barrier_first=85, barrier_last=70, period_months=3,
                 barriers_raw=[85, 85, 85, 85, 80, 80, 80, 80, 75, 75, 75, 70],
                 issue_date=date(2026, 4, 10), expiry_date=date(2029, 4, 11),
                 description=("[월지급식] 3년/3개월,25KI(85,85,85,85,80,80,80,80,75,75,75,70)%,"
                              "월수익행사율 65%,세전 연 19.8% (월 1.65%)"),
                 eval_dates=["2026-07-08", "2026-10-08", "2027-01-08", "2027-04-09",
                             "2027-07-09", "2027-10-08", "2028-01-07", "2028-04-07",
                             "2028-07-07", "2028-10-06", "2029-01-09", "2029-04-11"],
                 amount=26_230_000, status="조기상환"),
            dict(issuer="키움증권", product_no="1823", asset_type="종목형",
                 assets_raw="Tesla , Palantir", yield_rate=24.0, ki=30,
                 barrier_first=85, barrier_last=60, period_months=3,
                 barriers_raw=[85, 85, 80, 80, 75, 75, 75, 75, 70, 70, 60, 60],
                 issue_date=date(2026, 4, 10), expiry_date=date(2029, 4, 12),
                 description="[스텝다운] 3년/3개월/85-85-80-80-75-75-75-75-70-70-60-60/KI30",
                 eval_dates=["2026-07-09", "2026-10-09", "2027-01-08", "2027-04-09",
                             "2027-07-09", "2027-10-08", "2028-01-07", "2028-04-07",
                             "2028-07-07", "2028-10-09", "2029-01-09", "2029-04-12"],
                 amount=10_000_000, status="조기상환"),
            dict(issuer="키움증권", product_no="1826", asset_type="종목형",
                 assets_raw="Palantir   , Micron", yield_rate=31.71, ki=35,
                 barrier_first=75, barrier_last=60, period_months=4,
                 barriers_raw=[75, 75, 75, 70, 70, 60],
                 issue_date=date(2026, 4, 10), expiry_date=date(2028, 4, 13),
                 description="[스텝다운] 2년/4개월/75-75-75-70-70-60/KI35",
                 eval_dates=["2026-08-07", "2026-12-09", "2027-04-09", "2027-08-09",
                             "2027-12-09", "2028-04-13"],
                 amount=10_000_000, status="보유중"),
            dict(issuer="키움증권", product_no="1827", asset_type="종목형",
                 assets_raw="Palantir   , Micron", yield_rate=30.33, ki=35,
                 barrier_first=80, barrier_last=60, period_months=4,
                 barriers_raw=[80, 80, 75, 75, 70, 60],
                 issue_date=date(2026, 4, 10), expiry_date=date(2028, 4, 13),
                 description="[스텝다운리자드] 2년/4개월/80-80-75(L50)-75-70-60/KI35",
                 eval_dates=["2026-08-07", "2026-12-09", "2027-04-09", "2027-08-09",
                             "2027-12-09", "2028-04-13"],
                 amount=10_000_000, status="보유중"),
        ]
        for spec in self.rows:
            spec = dict(spec)
            amount, status = spec.pop("amount"), spec.pop("status")
            product = _product(**spec)
            Investment.objects.create(
                user=self.user, product=product, amount=amount, status=status,
                invested_at=product.issue_date)

    def test_six_rows_match_the_managers_sheet(self):
        expected = [
            ["삼성증권", "30868", "KOSPI200/S&P500/Nikkei225", 20260403, "29.04.02",
             20.0, 3000, "완료", 45, 90, 75, 3, "지수형", 202604, 202607, "", 150],
            ["키움증권", "1806", "Palantir/Micron", 20260403, "29.04.05",
             24.2, 3000, "", 20, 80, 65, 6, "종목형", 202604, 202610, "", 363],
            ["삼성증권", "30895", "현대차/삼성전자", 20260410, "29.04.11",
             19.8, 2623, "완료", 25, 85, 70, 3, "종목형", 202604, 202607, "월지급", 130],
            ["키움증권", "1823", "Tesla/Palantir", 20260410, "29.04.12",
             24.0, 1000, "완료", 30, 85, 60, 3, "종목형", 202604, 202607, "", 60],
            ["키움증권", "1826", "Palantir/Micron", 20260410, "28.04.13",
             31.71, 1000, "", 35, 75, 60, 4, "종목형", 202604, 202608, "", 106],
            ["키움증권", "1827", "Palantir/Micron", 20260410, "28.04.13",
             30.33, 1000, "", 35, 80, 60, 4, "종목형", 202604, 202608, "L50", 101],
        ]
        actual = pfx.build_rows(Investment.objects.select_related("product"))
        self.assertEqual(len(pfx.COLUMNS), 17)
        self.assertEqual(len(pfx.COLUMN_WIDTHS), 17)
        for got, want in zip(actual, expected):
            self.assertEqual(got, want, f"{want[0]} {want[1]} 행이 어긋난다")
        self.assertEqual(actual, expected)

    def test_expected_profit_is_reproducible_from_visible_columns(self):
        """예상수익이 같은 행의 투자금액·금리·주기만으로 재현되는가."""
        for row in pfx.build_rows(Investment.objects.select_related("product")):
            amount, rate, period, profit = row[6], row[5], row[11], row[16]
            self.assertEqual(profit, pfx._round_half_up(amount * rate / 100 * period / 12))


class ColumnRuleTest(TestCase):
    """개별 변환 규칙 — 단위·형식·폴백."""

    def setUp(self):
        self.user = get_user_model().objects.create_user(username="rule", password="x")

    def _inv(self, amount=10_000_000, status="보유중", **kw):
        kw.setdefault("issuer", "테스트증권")
        kw.setdefault("product_no", str(Product.objects.count() + 1))
        kw.setdefault("yield_rate", 12.0)
        kw.setdefault("barriers_raw", [90, 85, 80, 75])
        kw.setdefault("issue_date", date(2026, 4, 10))
        kw.setdefault("expiry_date", date(2027, 4, 10))
        kw.setdefault("period_months", 3)
        return Investment.objects.create(
            user=self.user, product=_product(**kw), amount=amount,
            status=status, invested_at=kw["issue_date"])

    def test_amount_is_in_manwon(self):
        self.assertEqual(pfx.build_row(self._inv(amount=30_000_000))[6], 3000)
        self.assertEqual(pfx.build_row(self._inv(amount=26_230_000))[6], 2623)
        self.assertEqual(pfx.build_row(self._inv(amount=2_000_000))[6], 200)

    def test_amount_keeps_fraction_when_not_a_round_manwon(self):
        self.assertEqual(pfx.build_row(self._inv(amount=10_005_000))[6], 1000.5)

    def test_date_formats(self):
        row = pfx.build_row(self._inv())
        self.assertEqual(row[3], 20260410)      # 발행일 YYYYMMDD 정수
        self.assertEqual(row[4], "27.04.10")    # 만기일 YY.MM.DD 문자열
        self.assertEqual(row[13], 202604)       # 투자월 YYYYMM 정수

    def test_real_issue_date_wins_over_issue_date(self):
        row = pfx.build_row(self._inv(real_issue_date=date(2026, 4, 13)))
        self.assertEqual(row[3], 20260413)
        self.assertEqual(row[13], 202604)

    def test_status_column(self):
        self.assertEqual(pfx.build_row(self._inv(status="보유중"))[7], "")
        self.assertEqual(pfx.build_row(self._inv(status="조기상환"))[7], "완료")
        # 손실 상환을 '완료'로 뭉뚱그리지 않는다
        self.assertEqual(pfx.build_row(self._inv(status="낙인후상환"))[7], "낙인후상환")
        self.assertEqual(pfx.build_row(self._inv(status="만기상환"))[7], "만기상환")

    def test_no_knock_in_label(self):
        self.assertEqual(pfx.build_row(self._inv(is_no_ki=True))[8], "노낙인")
        self.assertEqual(pfx.build_row(self._inv(ki=45))[8], 45)
        self.assertEqual(pfx.build_row(self._inv())[8], "")

    def test_first_eval_uses_confirmed_dates_when_available(self):
        inv = self._inv(barriers_raw=[90, 85, 80, 75],
                        eval_dates=["2026-07-08", "2026-10-08", "2027-01-08", "2027-04-09"])
        row = pfx.build_row(inv)
        self.assertEqual(row[14], 202607)
        self.assertNotIn("추정", row[15])

    def test_first_eval_falls_back_to_calculation_and_is_flagged(self):
        """eval_dates가 없어도 비우지 않는다 — 발행일 + 주기로 채우고 '추정'을 남긴다."""
        row = pfx.build_row(self._inv(eval_dates=None, period_months=4))
        self.assertEqual(row[14], 202608)
        self.assertEqual(row[15], "추정")

    def test_first_eval_respects_uneven_first_period(self):
        row = pfx.build_row(self._inv(eval_dates=None, period_months=3,
                                      first_eval_months=1))
        self.assertEqual(row[14], 202605)

    def test_first_eval_rolls_over_the_year(self):
        row = pfx.build_row(self._inv(eval_dates=None, issue_date=date(2026, 8, 7),
                                      period_months=6))
        self.assertEqual(row[14], 202702)

    def test_period_is_derived_when_missing(self):
        """주기가 비어 있으면 (만기-발행일) ÷ 배리어 개수로 유도한다."""
        row = pfx.build_row(self._inv(
            period_months=None, barriers_raw=[90, 85, 80, 75, 70, 65],
            issue_date=date(2026, 4, 10), expiry_date=date(2028, 4, 13)))
        self.assertEqual(row[11], 4)

    def test_blank_when_nothing_can_be_derived(self):
        """배리어도 주기도 없으면 추측하지 않고 비운다."""
        row = pfx.build_row(self._inv(period_months=None, barriers_raw=None,
                                      eval_dates=None, expiry_date=None))
        self.assertEqual(row[11], "")
        self.assertEqual(row[14], "")
        self.assertEqual(row[16], "")
        self.assertIn("확인필요", row[15])

    def test_note_combines_structure_flags_and_memo(self):
        inv = self._inv(description="[스텝다운리자드] 2년/4개월/80-80-75(L50)-75-70-60/KI35",
                        eval_dates=["2026-07-08", "2026-10-08", "2027-01-08", "2027-04-09"])
        inv.memo = "키움 CMA"
        inv.save()
        self.assertEqual(pfx.build_row(inv)[15], "L50 키움 CMA")

    def test_note_marks_lizard_without_a_marker(self):
        inv = self._inv(description="스텝다운 리자드형 상품",
                        eval_dates=["2026-07-08", "2026-10-08", "2027-01-08", "2027-04-09"])
        self.assertEqual(pfx.build_row(inv)[15], "리자드")

    def test_note_marks_monthly_payout_outside_the_bracket_tag(self):
        inv = self._inv(description="기초자산:현대차,SK하이닉스/3년만기 6개월단위 조기상환형/월지급식StepDown형",
                        eval_dates=["2026-07-08", "2026-10-08", "2027-01-08", "2027-04-09"])
        self.assertEqual(pfx.build_row(inv)[15], "월지급")

    def test_assets_keep_full_names(self):
        """축약하지 않는다 — 구분자만 정리한다."""
        row = pfx.build_row(self._inv(assets_raw="Palantir   , Micron"))
        self.assertEqual(row[2], "Palantir/Micron")
        row = pfx.build_row(self._inv(assets_raw="KOSPI200 , S&P500 , Nikkei225"))
        self.assertEqual(row[2], "KOSPI200/S&P500/Nikkei225")

    def test_rounding_is_half_up_like_excel(self):
        self.assertEqual(pfx._round_half_up(2.5), 3)
        self.assertEqual(pfx._round_half_up(129.8385), 130)
        self.assertEqual(pfx._round_half_up(105.7), 106)


class ExportIsolationTest(TestCase):
    """다른 사용자의 투자 기록이 파일에 섞이면 안 된다."""

    def setUp(self):
        User = get_user_model()
        self.me = User.objects.create_user(username="me", password="pw-me")
        self.other = User.objects.create_user(username="other", password="pw-other")
        self.mine = _product(
            issuer="삼성증권", product_no="30868", assets_raw="KOSPI200",
            yield_rate=20.0, ki=45, barrier_first=90, barrier_last=75,
            period_months=3, barriers_raw=[90, 85, 80, 75], asset_type="지수형",
            issue_date=date(2026, 4, 3), expiry_date=date(2027, 4, 2))
        self.theirs = _product(
            issuer="한국투자증권", product_no="99999", assets_raw="SECRETASSET",
            yield_rate=15.0, ki=50, barrier_first=90, barrier_last=70,
            period_months=6, barriers_raw=[90, 85, 80, 70], asset_type="종목형",
            issue_date=date(2026, 5, 1), expiry_date=date(2028, 5, 1))
        Investment.objects.create(user=self.me, product=self.mine,
                                  amount=30_000_000, invested_at=date(2026, 4, 3))
        Investment.objects.create(user=self.other, product=self.theirs,
                                  amount=77_770_000, invested_at=date(2026, 5, 1))

    def _download(self):
        self.client.login(username="me", password="pw-me")
        resp = self.client.get("/portfolio/export/")
        self.assertEqual(resp.status_code, 200)
        return resp

    def test_export_contains_only_my_rows(self):
        resp = self._download()
        book = openpyxl.load_workbook(io.BytesIO(resp.content))
        sheet = book["보유계약"]
        values = list(sheet.values)
        self.assertEqual(values[0], tuple(pfx.COLUMNS))
        self.assertEqual(len(values), 2, "본인 1건 외의 행이 있으면 안 된다")
        self.assertEqual(values[1][1], "30868")

    def test_other_users_data_is_absent_from_every_sheet(self):
        """시트 어디에도 남의 상품번호·자산명·금액이 나타나지 않아야 한다."""
        resp = self._download()
        book = openpyxl.load_workbook(io.BytesIO(resp.content))
        seen = [str(v) for ws in book.worksheets for row in ws.values for v in row]
        for leaked in ["99999", "SECRETASSET", "한국투자증권", "7777"]:
            self.assertNotIn(leaked, seen, f"{leaked} 가 새어 나갔다")

    def test_redeemed_rows_are_included(self):
        """보유중과 상환 완료분을 모두 담는다."""
        done = _product(
            issuer="키움증권", product_no="1823", assets_raw="Tesla , Palantir",
            yield_rate=24.0, ki=30, barrier_first=85, barrier_last=60,
            period_months=3, barriers_raw=[85, 80, 75, 60], asset_type="종목형",
            issue_date=date(2026, 4, 10), expiry_date=date(2029, 4, 12))
        Investment.objects.create(user=self.me, product=done, amount=10_000_000,
                                  invested_at=date(2026, 4, 10), status="조기상환",
                                  redeemed_at=date(2026, 7, 10),
                                  redeemed_amount=10_600_000)
        book = openpyxl.load_workbook(io.BytesIO(self._download().content))
        rows = list(book["보유계약"].values)[1:]
        self.assertEqual({r[1] for r in rows}, {"30868", "1823"})
        # 빈 문자열은 엑셀에서 빈 칸(None)으로 되읽힌다
        self.assertEqual({r[7] or "" for r in rows}, {"", "완료"})

    def test_anonymous_user_cannot_download(self):
        resp = self.client.get("/portfolio/export/")
        self.assertIn(resp.status_code, (301, 302))
        self.assertNotIn("30868", resp.content.decode("utf-8", "ignore"))

    def test_filename_is_utf8_encoded(self):
        resp = self._download()
        self.assertIn("filename*=UTF-8''", resp["Content-Disposition"])
