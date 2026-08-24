import calendar as pycalendar
import hmac
import logging
import math
from datetime import date, timedelta

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login as auth_login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.views import LoginView
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_POST

from . import ask_tools, portfolio_facts, telegram
from .asset_pages import TOP_ASSETS, asset_context
from .compare import MIN_PEERS, compare_context, peer_key, week_peer_counts
from .models import (
    Feedback, FeedbackBannerDismissal, ImportLog, Investment, Preset, Product,
    RedemptionVerdict, WatchItem, attach_peak_ratios, peak_ratios,
)

logger = logging.getLogger(__name__)

# 가족(운영진) 전용 — 공유 데이터(관심·프리셋·업로드)는 staff 계정만
family_required = user_passes_test(lambda u: u.is_active and u.is_staff, login_url="/accounts/login/")

# 운영자(superuser) 전용 — 회원 관리
admin_required = user_passes_test(lambda u: u.is_active and u.is_superuser, login_url="/accounts/login/")


def _scope(qs, user):
    """프리셋/관심 소유 범위 — 계정별 완전 분리.

    (구버전은 staff끼리 user=None '공용'을 공유했으나, 공개 서비스로 전환하면서
     운영자 계정끼리 관심목록·프리셋이 섞이는 문제가 있어 개인화로 통일했다.)
    """
    if not user.is_authenticated:
        return qs.none()
    return qs.filter(user=user)


from django import forms as _forms


class SignUpForm(UserCreationForm):
    """가입 폼 — 이메일 필수 (아이디/비밀번호 찾기에 사용)."""
    email = _forms.EmailField(required=True)

    def save(self, commit=True):
        user = super().save(commit=False)
        user.email = self.cleaned_data["email"]
        if commit:
            user.save()
        return user


class RememberLoginView(LoginView):
    """로그인 유지 체크 시 30일, 미체크 시 브라우저 종료로 세션 만료."""
    template_name = "core/login.html"

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.POST.get("remember"):
            self.request.session.set_expiry(60 * 60 * 24 * 30)  # 30일
        else:
            self.request.session.set_expiry(0)  # 브라우저 닫으면 로그아웃
        return response


def signup(request):
    """회원가입 — 일반 회원은 공개 화면 + 본인 포트폴리오만 사용."""
    if request.user.is_authenticated:
        return redirect("weekly")
    if request.method == "POST":
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            auth_login(request, user)
            messages.success(request, "가입을 환영합니다! 포트폴리오에서 투자를 등록해보세요.")
            return redirect("weekly")
    else:
        form = SignUpForm()
    return render(request, "core/signup.html", {"form": form})


# ── 소셜 로그인 (카카오·구글) ──────────────────────────────
# allauth 의 provider 뷰 앞에 세우는 관문. 같은 경로에 우리 것을 먼저 등록해
# 두면 URL 해석은 여기로 오고, {% url 'kakao_login' %} 역참조는 allauth 가
# 등록한 이름 그대로 쓸 수 있다(경로가 같으므로 결과도 같다).
#
# 왜 필요한가: 키(.env 의 CLIENT_ID)가 없으면 allauth 는 소셜 앱을 못 찾아
# 예외를 던진다. 아직 키를 발급받기 전이거나 배포 때 .env 를 빠뜨리면
# 방문자가 에러 화면을 보게 된다. 그럴 땐 조용히 로그인 화면으로 되돌린다.
#
# provider별로 개별 확인한다(전체 SOCIAL_LOGIN_ENABLED가 아니다) — 카카오만
# 키가 있고 구글은 아직 없는 상태에서 구글 진입점이 "하나라도 켜져 있으니
# 통과"로 잘못 판단해 allauth로 넘겼다가 그대로 500이 나는 걸 막기 위해서다.
def _social_disabled_redirect(request, label):
    messages.error(request, f"{label} 로그인은 현재 준비 중입니다. 아이디로 로그인해 주세요.")
    return redirect("login")


def _provider_enabled(provider_id):
    return getattr(settings, "SOCIAL_PROVIDERS", {}).get(provider_id, {}).get("enabled", False)


def kakao_login_entry(request):
    """/accounts/kakao/login/ — 키가 있을 때만 allauth 로 넘긴다."""
    if not _provider_enabled("kakao"):
        return _social_disabled_redirect(request, "카카오")
    from allauth.socialaccount.providers.kakao.views import oauth2_login
    return oauth2_login(request)


def kakao_callback_entry(request):
    """/accounts/kakao/login/callback/ — 위와 같은 이유의 관문."""
    if not _provider_enabled("kakao"):
        return _social_disabled_redirect(request, "카카오")
    from allauth.socialaccount.providers.kakao.views import oauth2_callback
    return oauth2_callback(request)


def google_login_entry(request):
    """/accounts/google/login/ — 키가 있을 때만 allauth 로 넘긴다."""
    if not _provider_enabled("google"):
        return _social_disabled_redirect(request, "구글")
    from allauth.socialaccount.providers.google.views import oauth2_login
    return oauth2_login(request)


def google_callback_entry(request):
    """/accounts/google/login/callback/ — 위와 같은 이유의 관문."""
    if not _provider_enabled("google"):
        return _social_disabled_redirect(request, "구글")
    from allauth.socialaccount.providers.google.views import oauth2_callback
    return oauth2_callback(request)


def naver_login_entry(request):
    """/accounts/naver/login/ — 키가 있을 때만 allauth 로 넘긴다."""
    if not _provider_enabled("naver"):
        return _social_disabled_redirect(request, "네이버")
    from allauth.socialaccount.providers.naver.views import oauth2_login
    return oauth2_login(request)


def naver_callback_entry(request):
    """/accounts/naver/login/callback/ — 위와 같은 이유의 관문."""
    if not _provider_enabled("naver"):
        return _social_disabled_redirect(request, "네이버")
    from allauth.socialaccount.providers.naver.views import oauth2_callback
    return oauth2_callback(request)


def social_connect_help(request):
    """자동 연결을 일부러 하지 않았을 때 보여주는 안내 화면.

    이 화면이 뜨는 두 경우 (core/socialauth.py 참조)
      · 카카오가 준 이메일을 카카오가 검증하지 않았는데, 그 이메일을 쓰는
        기존 계정이 있다 → 그대로 붙이면 남의 계정을 가져가는 길이 된다.
      · 카카오에서 이메일을 아예 못 받았다 → 기존 회원인지 판별할 수 없다.
    """
    reason = request.GET.get("reason", "")
    if reason not in ("unverified", "no-email"):
        reason = "unverified"
    return render(request, "core/social_connect_help.html", {"reason": reason})


def find_id(request):
    """아이디 찾기 — 가입 이메일 입력 시 마스킹된 아이디 표시."""
    from django.contrib.auth import get_user_model
    found = None
    searched = False
    if request.method == "POST":
        searched = True
        email = request.POST.get("email", "").strip()
        if email:
            names = list(get_user_model().objects.filter(email__iexact=email)
                         .values_list("username", flat=True))
            found = [n[:2] + "*" * max(len(n) - 4, 1) + n[-2:] if len(n) > 4
                     else n[0] + "*" * (len(n) - 1) for n in names]
    return render(request, "core/find_id.html", {"found": found, "searched": searched})


def _week_range(offset: int = 0):
    """offset주 뒤의 (월요일, 일요일)."""
    today = date.today()
    monday = today - timedelta(days=today.weekday()) + timedelta(weeks=offset)
    return monday, monday + timedelta(days=6)


# ── 주간 청약 (메인) ──────────────────────────────
WEEKLY_FILTER_PARAMS = ["asset", "ki_max", "yield_min", "currency",
                        "no_ki", "issuer", "preset", "sort", "dir"]

# 주 이동 폭의 한계(약 100년). 이보다 크면 timedelta(weeks=offset)가 date 범위를
# 벗어나 OverflowError를 낸다 — ?w=999999가 그랬다.
MAX_WEEK_OFFSET = 5200


def _int_param(raw, default=None, lo=None, hi=None):
    """GET으로 들어온 문자열을 정수로. 무효하거나 범위 밖이면 default.

    화면에 오류를 띄우지 않는다. 주소창을 손댔거나 링크가 깨진 것뿐이라
    방문자가 할 수 있는 일이 없고, ?preset=abc를 공란으로 되돌린 것과 같은 처리다.
    """
    try:
        value = int(raw)
    except (TypeError, ValueError):
        return default
    if (lo is not None and value < lo) or (hi is not None and value > hi):
        return default
    return value


def _float_param(raw, default=None):
    """같은 규칙의 실수판. nan·inf는 비교가 성립하지 않으므로 무효로 본다."""
    try:
        value = float(raw)
    except (TypeError, ValueError):
        return default
    return value if math.isfinite(value) else default


# 시장 국면 카드에 표시할 대표 지수
#
# HSCEI(홍콩H)는 2026-08-18에 추가했다. 카드를 처음 만든 8/2에 빠져 있었는데
# 이유가 코드에도 커밋에도 없었다 — 기술적 제약이 아니라 목록 누락이었다.
# 넣는 근거: ① 최근 12주 청약에도 114건이 이 지수를 쓴다 ② 10년 실측에서
# 원금손실 3,220건 중 2,245건(69.7%)에 이 지수가 들어 있었다. ELS 위험을
# 읽는 지표판에서 이 지수가 빠지면 정작 가장 중요한 칸이 비는 셈이다.
REGIME_INDEXES = [("KOSPI200", "KOSPI200 Index"), ("S&P500", "S&P500 Index"),
                  ("Nikkei225", "Nikkei225 Index"), ("Euro Stoxx 50", "Euro Stoxx 50 Index"),
                  ("HSCEI", "HSCEI Index")]
# 등급(양호/주의/과열) 라벨은 붙이지 않는다 — 분기 발행 기준으로 만든 눈금과
# 주간 청약 기준 통과율은 모수가 달라 단정할 근거가 없다. 사실만 전달한다.
# (2026-08-03 태훈님: "굳이 평가를 우리가 할 필요는 없다")


def _market_regime(monday, sunday):
    """시장 국면 — 이번 주 조건 통과율 + 대표 지수 위치.

    ▣ 통과율(rate)은 그 주 상품이 있어야 성립하지만 지수 위치(indexes)는 아니다.
      상품이 0건이면 rate=None으로 두고 지수 위치만 채워 돌려준다. 예전엔 여기서
      통째로 return None이라, 청약 0건인 주(2026-08-17 주부터 실제로 발생)에는
      시장 국면 카드가 화면에서 사라졌다 — 지수 위치는 그 주에 상품이 없어도
      유효한 정보라 남겨야 한다. 종목형 기초자산 표(stocks)는 그 주 상품에서
      뽑는 것이라 상품이 없으면 비는 게 맞다. (2026-08-18)

    통과율은 시세가 필요 없는 두 조건(직전 연도 낙인 컷 · 1차 배리어)으로만 센다.
    10년 분기 검증에서 통과율 ↔ 이후 1년 주식수익률 상관 +0.52 — 즉 조건이
    나쁜 시기는 시장도 과열된 시기였다. '갈아타라'가 아니라 '지금이 어디인가'를
    읽는 용도. (regime_signal.py 검증, 2026-08-02)

    ⚠ 여기 쓰는 컷(v7_ki_cut)은 **배지 게이트가 아니다.** v8부터 배지는 그 주차
    그룹 내부 분포에서 컷을 잡는다(models._compute_radar_pool). 이 계기판만
    직전 연도 절대 컷으로 남긴 것은 의도한 것이다 — 그룹 상대 컷으로 바꾸면
    통과율이 정의상 늘 40% 근처가 돼 국면을 전혀 못 읽는다. (2026-08-11)
    """
    from core import market as _m
    from core.models import (
        RADAR_V7_B0_MAX, _peak_fetch_days, _peak_from_series, v7_ki_cut,
    )

    # 원금지급형(ELB·DLB) 제외 — 낙인이 없어 통과(n_pass)에는 절대 못 들어가면서
    # 모수(n_all)만 키우고 있었다. 그래서 통과율이 실제보다 낮게 찍혔다.
    # (2026-08-06 실측: 이번 주 109/266=41.0% → 109/219=49.8%)
    group = list(Product.objects.listed().filter(sub_end__gte=monday, sub_end__lte=sunday))
    n_all = n_pass = 0
    for p in group:
        if p.asset_type not in RADAR_V7_B0_MAX:
            continue
        n_all += 1
        cut = v7_ki_cut(p.asset_type)
        # 낙인 비교를 "이하"(<=)로 쓴다 — 2026-08-11 조 팀장 지시.
        # ⚠ v7_ki_cut이 실제 배지 선정에 쓰이던 시절(v8 이전)엔 "미만"(<)이
        # 필수였다 — "이하"면 컷과 정확히 같은 낙인 상품이 새어 들어와
        # 2018·2021년 손실로 이어진 이력이 있다(els-radar-badge-constants
        # 검증). v8부터 배지 선정은 이 함수를 아예 안 쓰고 그 주차 그룹
        # 상대 컷(models.group_cut)을 쓰므로 그 위험은 더는 안 걸린다 —
        # v7_ki_cut은 지금 이 국면 계기판에서만 참고용으로 남아 있다.
        if (not p.is_no_ki and p.ki is not None and cut is not None and p.ki <= cut
                and p.barrier_first is not None
                and p.barrier_first <= RADAR_V7_B0_MAX[p.asset_type]):
            n_pass += 1
    # 모수가 0이면 통과율을 낼 수 없다 — 0%가 아니라 '없음'이다. 화면도 이때는
    # 통과율 칸을 통째로 감춘다(weekly.html).
    rate = round(n_pass / n_all * 100, 1) if n_all else None

    def _gauge(label, name):
        # 화면 표기가 "52주 고점 대비 / 직전 1년 대비"이므로 창도 캘린더 52주로 자른다.
        # (예전엔 fetch_history(days=370)의 종가 전체를 썼는데 370은 거래일 수로
        #  해석돼 실제로는 1.5년 창이었다 — 실측 557캘린더일. 2026-08-05)
        tk = _m.resolve_ticker(name)
        hist = _m.fetch_history(tk, days=_peak_fetch_days(date.today())) if tk else None
        if not hist:
            return None
        peak, ret1y = _peak_from_series(hist, hist[-1][0])
        if peak is None or ret1y is None:
            return None      # 표에 두 값이 다 있어야 한 행이 성립한다
        return {"name": label, "peak": round(peak, 1), "ret1y": round(ret1y, 1)}

    idx = [g for g in (_gauge(lb, nm) for lb, nm in REGIME_INDEXES) if g]

    # 종목형 기초자산 — 이번 주 종목형 상품에 가장 많이 등장하는 5개
    freq = {}
    for p in group:
        if p.asset_type != "종목형":
            continue
        for a in _m.split_assets(p.assets_raw or ""):
            a = a.strip()
            if a:
                freq[a] = freq.get(a, 0) + 1
    # 지수는 위 표에 이미 있으므로 제외 (혼합형 상품 때문에 섞여 들어온다)
    idx_tickers = {_m.resolve_ticker(nm) for _, nm in REGIME_INDEXES}
    stocks = []
    for name, cnt in sorted(freq.items(), key=lambda x: -x[1]):
        if _m.resolve_ticker(name) in idx_tickers:
            continue
        g = _gauge(_m.shorten_asset_display(name), name)
        if g:
            g["cnt"] = cnt
            stocks.append(g)
        if len(stocks) >= 5:
            break

    # 통과 조건을 화면에 그대로 노출 (컷은 매년 자동 산출되므로 하드코딩 금지).
    # 유형별로 줄을 나눈다 — 한 줄로 붙여 쓰면 두 유형 조건이 섞여 읽힌다
    # (2026-08-11 조 팀장 지시). 낙인·1차 조기상환 둘 다 "이하"로 통일 —
    # 위 판정 로직과 맞춘 것이다.
    cond = [
        f"{t} 낙인 {v7_ki_cut(t)} 이하 · 1차 조기상환 {RADAR_V7_B0_MAX[t]} 이하"
        for t in ("지수형", "종목형")
    ]
    # 보여줄 게 하나도 없을 때만 카드를 통째로 접는다 — 통과율도 못 내고
    # 지수 시세도 전부 못 읽은 경우다.
    if rate is None and not idx:
        return None
    return {"n_all": n_all, "n_pass": n_pass, "rate": rate,
            "cond": cond, "indexes": idx, "stocks": stocks}


# 이 일수를 넘겨 수집 기록이 없으면 '수집이 멈췄다'로 본다(배지 경고색).
FRESHNESS_STALE_DAYS = 7


def _freshness(last_import):
    """수집 신선도 배지 — '수집이 멈춘 것'과 '수집할 청약이 없는 것'을 나눈다.

    ▣ 왜 나눠야 하나 (2026-08-18)
      배지는 마지막 ImportLog 시각만 보고 "N일 전 데이터"라고 썼다. 그런데
      scrape_kofia는 **수집된 게 0건이면 ImportLog를 아예 안 남기고** 돌아갔다.
      그래서 청약이 한 건도 안 나오는 기간에는 배치가 매일 정상 실행되는데도
      배지가 마지막으로 상품이 있던 날에 멈춰, 방문자에게는 '서비스가 죽었다'로
      읽혔다. 배치 쪽에서 0건 실행도 기록으로 남기게 고쳤고(row_count=0),
      여기서는 그 기록을 'quiet'로 구분해 문구를 달리 쓴다.

    상태
      fresh  마지막 수집에서 청약 상품을 받아왔다 — 평소 상태
      quiet  배치는 돌았는데 그 시점에 청약 중인 상품이 0건이었다
      stale  수집 기록 자체가 FRESHNESS_STALE_DAYS일 넘게 없다 — 진짜 경고

    ⚠ 순서가 중요하다. 0건 실행이 일주일 넘게 이어지면 그건 수집이 멈춘 것과
      구분이 안 되므로 stale이 이긴다.
    """
    if not last_import:
        return None
    days = (date.today() - last_import.imported_at.date()).days
    if days > FRESHNESS_STALE_DAYS:
        state = "stale"
    elif last_import.row_count == 0:
        state = "quiet"
    else:
        state = "fresh"
    return {"days": days, "state": state, "rows": last_import.row_count}


def _empty_week_guide(monday, offset):
    """청약 마감 상품이 한 건도 없는 주에 대신 보여줄 곳을 찾는다.

    ⚠ **왜 없는지는 계산하지 않는다.** KOFIA는 '청약 중 0건'만 알려줄 뿐 사유를
      싣지 않으므로 우리가 아는 사실이 아니다. 화면 문구도 사유를 단정하지
      않는다(weekly.html의 빈 주 안내 참고).

    돌려주는 것은 '가장 가까운, 상품이 있는 주' 하나뿐이다. 과거 쪽을 먼저 보고
    (직전 주가 방문자에게 제일 쓸모 있다), 없으면 미래 쪽을 본다 — 서비스 데이터
    시작보다 앞선 주(?w=-500 같은 링크)로 들어오면 과거 쪽이 비기 때문이다.
    """
    qs = Product.objects.listed()
    sunday = monday + timedelta(days=6)
    ref = (qs.filter(sub_end__lt=monday).order_by("-sub_end")
           .values_list("sub_end", flat=True).first())
    if ref is None:
        ref = (qs.filter(sub_end__gt=sunday).order_by("sub_end")
               .values_list("sub_end", flat=True).first())
    if ref is None:
        return {"near": None}
    near_monday = ref - timedelta(days=ref.weekday())
    delta = (near_monday - monday).days // 7
    return {"near": {
        "monday": near_monday,
        "sunday": near_monday + timedelta(days=6),
        "offset": offset + delta,
        "weeks": abs(delta),
        "past": delta < 0,
        "count": qs.filter(sub_end__gte=near_monday,
                           sub_end__lte=near_monday + timedelta(days=6)).count(),
    }}


def weekly(request):
    # ── 필터 세션 저장/복원 ──
    if "reset" in request.GET:
        request.session.pop("weekly_filters", None)
        return redirect("weekly")

    # 빈 URL(메뉴 클릭)로 오면 저장된 필터 복원
    if not request.GET and request.session.get("weekly_filters"):
        from django.urls import reverse
        return redirect(reverse("weekly") + "?" + request.session["weekly_filters"])

    # 그 외에는 현재 필터 상태를 저장 (주 이동 파라미터 w 제외)
    _saved = request.GET.copy()
    _saved.pop("w", None)
    request.session["weekly_filters"] = _saved.urlencode()

    # 주차 이동 w도 ?preset과 같은 유형이었다 — GET 문자열을 그대로 int()에 넣어
    # 무효한 값이면 /weekly/가 통째로 500이었다. (2026-08-07 실측:
    # ?w=abc·?w=1.5·?w=·?w=1e3·?w=0x10 → ValueError, ?w=999999 → timedelta의
    # OverflowError.) 무효한 값은 조용히 이번 주(0)로 되돌린다.
    offset = _int_param(request.GET.get("w"), 0, lo=-MAX_WEEK_OFFSET, hi=MAX_WEEK_OFFSET)
    monday, sunday = _week_range(offset)

    qs = Product.objects.listed().filter(sub_end__gte=monday, sub_end__lte=sunday)

    # 필터
    f_asset = request.GET.get("asset", "")
    f_ki_max = request.GET.get("ki_max", "")
    f_yield_min = request.GET.get("yield_min", "")
    f_currency = request.GET.get("currency", "")
    f_no_ki = request.GET.get("no_ki", "")
    f_issuers = request.GET.getlist("issuer")  # 다중 선택
    # 프리셋은 '각자 자기 조건' — 본인 것만 걸린다.
    # 예전엔 Preset.objects.get(id=...)라 소유자를 안 봤다. 비로그인 방문자가
    # ?preset=1 하나로 남의 프리셋 조건을 그대로 적용받았고(2026-08-06 실측),
    # 화면에는 그 프리셋 칩이 없으니 왜 걸러졌는지 알 길도 없었다.
    # 숫자가 아닌 값(?preset=abc)은 int 변환에서 ValueError가 나 500이 됐다.
    # 남의 것·없는 것·숫자 아닌 것은 모두 공란으로 되돌린다 — '전체'가 켜진다.
    preset_id = request.GET.get("preset", "")
    preset = None
    if preset_id.isdigit():
        preset = _scope(Preset.objects.all(), request.user).filter(id=preset_id).first()
    if preset is None:
        preset_id = ""

    # 숫자 칸(낙인 이하 · 수익률 이상)도 같은 자리였다 — ?ki_max=abc·?ki_max=45.5·
    # ?yield_min=abc가 int()·float()에서 ValueError를 내 500이었다(2026-08-07 실측).
    # 무효하면 그 칸을 안 건 것으로 보고, 화면 입력칸도 공란으로 되돌린다.
    v_ki_max = _int_param(f_ki_max)
    if v_ki_max is None:
        f_ki_max = ""
    v_yield_min = _float_param(f_yield_min)
    if v_yield_min is None:
        f_yield_min = ""

    if preset:
        qs = preset.match_queryset(qs)
    else:
        if f_issuers:
            qs = qs.filter(issuer__in=f_issuers)
        if f_asset:
            qs = qs.filter(asset_type=f_asset)
        if f_currency:
            qs = qs.filter(currency=f_currency)
        if v_yield_min is not None:
            qs = qs.filter(yield_rate__gte=v_yield_min)
        if v_ki_max is not None:
            from django.db.models import Q
            cond = Q(is_no_ki=False, ki__lte=v_ki_max)
            if f_no_ki != "exclude":
                cond |= Q(is_no_ki=True)
            qs = qs.filter(cond)

    # ── 정렬 ──────────────────────────────────────
    from django.db.models import F
    SORT_FIELDS = {
        "issuer": "issuer", "product_no": "product_no", "assets": "assets_raw",
        "yield": "yield_rate", "ki": "ki", "first": "barrier_first",
        "last": "barrier_last", "period": "period_months", "type": "asset_type",
        "sub_end": "sub_end", "loss": "loss_prob",
    }
    # term_months·peak_ratio는 계산값(DB 컬럼 아님) → Python 정렬
    PY_SORT = {"term": lambda p: p.term_months, "peak": lambda p: p.peak_ratio}

    # 기본 정렬: 수익률 내림차순 (정렬 파라미터 없을 때)
    sort_key = request.GET.get("sort", "yield")
    if sort_key not in PY_SORT and sort_key not in SORT_FIELDS:
        sort_key = "yield"
    sort_dir = request.GET.get("dir", "desc" if "sort" not in request.GET else "asc")
    if sort_key in PY_SORT:
        products = list(qs.order_by("-yield_rate"))
    else:
        field = SORT_FIELDS[sort_key]
        ordering = (F(field).desc(nulls_last=True) if sort_dir == "desc"
                    else F(field).asc(nulls_last=True))
        products = list(qs.order_by(ordering, "-yield_rate"))

    # 고점대비 — 표시·정렬 모두에 필요하므로 정렬 전에 붙인다
    attach_peak_ratios(products)

    if sort_key in PY_SORT:
        # None은 정렬 방향과 무관하게 항상 뒤로 (sentinel)
        desc = sort_dir == "desc"
        sentinel = float("-inf") if desc else float("inf")
        getter = PY_SORT[sort_key]
        products.sort(
            key=lambda p: getter(p) if getter(p) is not None else sentinel,
            reverse=desc,
        )

    # ── 유사상품 비교 버튼 라벨 ──────────────────
    # 같은 주 · 같은 자산유형 · 같은 낙인값 상품이 몇 건인지. 상품마다 세면 200번
    # 넘게 왕복하므로 한 주 전체를 값만 한 번 긁어 센다(쿼리 1회).
    # 필터와 무관하게 그 주 전체가 모수다 — 내가 화면에서 좁혀 놓은 것 때문에
    # 비교 대상이 줄면 백분위가 사용자 필터마다 달라져 뜻이 없어진다.
    # 3건 미만이면 0으로 둬서 버튼 자체를 감춘다(혼자 1등인 건 의미가 없다).
    _peer_counts = week_peer_counts(monday, sunday)
    for p in products:
        n = _peer_counts.get(peer_key(p), 0)
        p.compare_n = n if n >= MIN_PEERS else 0

    # 정렬 헤더용 컬럼 메타 (URL은 현재 필터 유지 + 정렬 토글)
    base_params = request.GET.copy()
    base_params.pop("sort", None)
    base_params.pop("dir", None)

    def _sort_url(key):
        p = base_params.copy()
        p["sort"] = key
        p["dir"] = "desc" if (sort_key == key and sort_dir == "asc") else "asc"
        return "?" + p.urlencode()

    col_defs = [
        ("issuer", "발행사", False), ("product_no", "상품번호", False),
        ("assets", "기초자산", False), ("yield", "수익률", True),
        ("ki", "낙인", True), ("first", "1차", True), ("last", "막차", True),
        ("term", "기간", True), ("period", "주기", True),
        ("peak", "고점대비", True), ("loss", "손실확률", True),
        ("type", "유형", False), ("sub_end", "마감", True),
    ]
    columns = [
        {"key": k, "label": lbl, "num": num, "url": _sort_url(k),
         "active": sort_key == k, "dir": sort_dir}
        for k, lbl, num in col_defs
    ]

    watched_ids = set(_scope(WatchItem.objects.all(), request.user).values_list("product_id", flat=True))
    invested_ids = set()
    if request.user.is_authenticated:
        invested_ids = set(Investment.objects.filter(
            user=request.user, status="보유중").values_list("product_id", flat=True))
    freshness = _freshness(ImportLog.objects.first())

    regime = _market_regime(monday, sunday)

    # 발행사 필터 후보 (이번 주 상품에 존재하는 발행사)
    # 목록과 같은 .listed()를 태운다 — 안 그러면 골라도 0건인 발행사가 남는다
    # (이번 주 17곳 중 3곳은 ELB·DLB만 낸 발행사였다. 2026-08-06 실측)
    issuers = sorted(set(
        Product.objects.listed().filter(sub_end__gte=monday, sub_end__lte=sunday)
        .values_list("issuer", flat=True)
    ))

    # ── 청약이 한 건도 없는 주 ────────────────────
    # '필터를 좁혀서 0건'과 '그 주에 아예 상품이 없어서 0건'은 다른 상황인데
    # 예전엔 둘 다 "이번 주 청약 마감 상품이 없습니다" 한 줄로 끝났다.
    # 세는 쿼리는 목록이 비었을 때만 돈다(평소 주에는 추가 비용 없음).
    empty_week = None
    if not products:
        week_total = Product.objects.listed().filter(
            sub_end__gte=monday, sub_end__lte=sunday).count()
        if not week_total:
            empty_week = _empty_week_guide(monday, offset)

    # ── 이번주 TOP5 (현재 주만) ──
    # 아주 강한 신호 상품 중 손실확률 0% & 1년내 조기상환 ≥90% → 수익률 상위 5.
    # (고쿠폰이지만 손실확률이 있는 상품은 '강한 신호'에서 확인)
    # 중복도 = 보유 포트폴리오 중 같은 기초자산을 가진 투자금 비중.
    # 선정 로직은 models.radar_top5()로 통일(주간요약 텔레그램과 동일 기준).
    # overlap(보유 중복도) 계산만 뷰에 남긴다.
    # ── v7 TOP5 이원화: 안정(지수형) / 수익(종목형) 트랙 ──
    top5_tracks = []
    if offset >= 0:  # 지난 주 조회 시에는 표시 안 함
        from core import market as _mkt
        from core.models import radar_tracks

        def _asset_keys(raw):
            return {_mkt.resolve_ticker(a) or a for a in _mkt.split_assets(raw)}

        # 본인 보유 기준 — 전 회원 합산이면 남의 포트폴리오와의 중복률이 나온다.
        # 익명 사용자에게는 조회 자체를 하지 않는다. user=AnonymousUser로 filter를
        # 걸면 DB에 닿기도 전에 TypeError가 나서 /weekly/가 통째로 500이 됐다
        # (2026-08-03, 비로그인 방문자 전원 영향. 어차피 결과는 staff에게만 쓴다).
        inv_assets = []
        if request.user.is_authenticated:
            inv_assets = [
                (inv.amount, _asset_keys(inv.product.assets_raw))
                for inv in Investment.objects.filter(
                    user=request.user, status="보유중").select_related("product")
            ]
        total_held = sum(amt for amt, _ in inv_assets)

        # 중복도는 가족(staff) 계정에만 표시 — 외부인에게 보유 성향 노출 방지
        show_overlap = request.user.is_authenticated and request.user.is_staff
        tracks = radar_tracks(monday, sunday)
        TRACK_META = {
            "지수형": {"label": "지수형 TOP5",
                      "sub": "1차 조기상환 90 이하 · 고점 회피(완만한 상승은 예외) · "
                             "낙인이 가장 낮은 40% 미만", "icon": "fa-shield-halved"},
            "종목형": {"label": "종목형 TOP5",
                      "sub": "1차 조기상환 80 이하 · 고점 회피(완만한 상승은 예외) · "
                             "낙인이 가장 낮은 40% 미만", "icon": "fa-rocket"},
        }
        for tier in ("지수형", "종목형"):
            items = []
            for p in tracks.get(tier, []):
                overlap_pct = None
                if show_overlap:
                    pkeys = _asset_keys(p.assets_raw)
                    overlap = sum(amt for amt, keys in inv_assets if keys & pkeys)
                    overlap_pct = round(overlap / total_held * 100) if total_held else 0
                r = p.radar or {}
                items.append({"p": p, "overlap_pct": overlap_pct,
                              "peak": r.get("peak")})
            if items:
                meta = TRACK_META[tier]
                top5_tracks.append({
                    "tier": tier, "label": meta["label"], "sub": meta["sub"],
                    "icon": meta["icon"],
                    "color": items[0]["p"].radar["color"], "items": items,
                })

    # ── 낙인대별 최고 수익 (선택한 주차 상품 · 유형 탭 × 낙인 값별 수익률 top5) ──
    # 대상 낙인 값은 태훈님 지정: 종목형 15~30, 지수형 25~40. 노낙인 제외.
    # TOP5처럼 주차를 따라간다 (주 이동 시 그 주의 분포로 갱신, 필터바 무관).
    # 지수형 먼저 (탭 기본 선택도 지수형) — 2026-08-03 태훈님 요청
    KI_BUCKETS = {"지수형": (25, 30, 35, 40), "종목형": (15, 20, 25, 30)}
    from collections import defaultdict
    _buckets = defaultdict(list)
    for p in Product.objects.listed().filter(
            sub_end__gte=monday, sub_end__lte=sunday,
            is_no_ki=False, ki__isnull=False):
        t = p.asset_type
        if t in KI_BUCKETS and p.ki in KI_BUCKETS[t]:
            _buckets[(t, int(p.ki))].append(p)
    ki_top5 = []
    for t, kis in KI_BUCKETS.items():
        cols = []
        for k in kis:
            lst = _buckets.get((t, k), [])
            if not lst:
                continue
            top = sorted(lst, key=lambda p: p.yield_rate or 0, reverse=True)[:5]
            cols.append({"ki": k, "count": len(lst), "top": top})
        if cols:
            ki_top5.append({"type": t, "cols": cols})
    if not ki_top5:
        ki_top5 = None

    return render(request, "core/weekly.html", {
        "meta_desc": ("이번 주 청약하는 ELS를 한 곳에 모아 낙인 배리어·연 수익률·"
                      "만기손실확률·조기상환 조건까지 한 표로 비교합니다. 매주 월요일 갱신."),
        "regime": regime,
        "products": products,
        "top5_tracks": top5_tracks,
        "ki_top5": ki_top5,
        "columns": columns,
        "monday": monday, "sunday": sunday, "offset": offset,
        "total": len(products),
        "presets": _scope(Preset.objects.all(), request.user),
        "issuers": issuers,
        "watched_ids": watched_ids,
        "invested_ids": invested_ids,
        "freshness": freshness,
        "empty_week": empty_week,
        "filters": {
            "asset": f_asset, "ki_max": f_ki_max, "yield_min": f_yield_min,
            "currency": f_currency, "no_ki": f_no_ki, "preset": preset_id,
            "issuers": f_issuers,
        },
        "has_saved_filters": bool(request.session.get("weekly_filters")),
        "show_feedback_banner": should_show_feedback_banner(request.user),
        "active_nav": "weekly",
    })


# ── 상품 상세 ─────────────────────────────────────
def _compare_html(request, product, same_assets=False):
    """비교 패널을 미리 그려 둔 HTML. 비교가 성립하지 않으면 None."""
    ctx = compare_context(product, same_assets=same_assets)
    if ctx is None:
        return None
    return render_to_string("core/_compare_panel.html", ctx, request=request)


def _product_meta_desc(product):
    """상품 상세의 검색 결과 스니펫.

    상품마다 값이 달라야 의미가 있다 — 같은 문장이 3천 개 붙으면 검색엔진이
    전부 같은 페이지로 본다. 그래서 실제로 채워진 값만 골라 붙인다.
    """
    head = " ".join(x for x in (product.issuer, product.product_no) if x)
    facts = []
    if product.assets_raw:
        facts.append(f"기초자산 {product.assets_raw}")
    if product.yield_rate is not None:
        facts.append(f"연 수익률 {product.yield_rate:g}%")
    if product.is_no_ki:
        facts.append("노낙인")
    elif product.ki is not None:
        facts.append(f"낙인 {product.ki}%")
    if product.loss_prob is not None:
        facts.append(f"만기손실확률 {product.loss_prob:g}%")
    body = ", ".join(facts)
    tail = "조기상환 배리어와 평가일, 낙인까지 남은 거리를 한눈에 확인하세요."
    return f"{head} — {body}. {tail}" if body else f"{head} 상품 조건 상세. {tail}"


def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    is_watched = _scope(WatchItem.objects.filter(product=product), request.user).exists()

    # 배리어 계단 SVG용 데이터
    barriers = product.barriers_raw or []
    svg = None
    if barriers:
        w, h, pad = 640, 220, 36
        n = len(barriers)
        bmin, bmax = min(barriers), max(barriers)
        span = max(bmax - bmin, 10)
        steps = []
        step_w = (w - pad * 2) / n
        for i, b in enumerate(barriers):
            x = pad + i * step_w
            y = pad + (bmax - b) / span * (h - pad * 2)
            steps.append({
                "x": round(x, 1), "x2": round(x + step_w - 4, 1),
                "cx": round(x + step_w / 2, 1),
                "y": round(y, 1), "v": b, "n": i + 1,
            })
        ki_y = None
        if product.ki is not None:
            ki_y = pad + (bmax - product.ki) / span * (h - pad * 2)
            ki_y = min(ki_y, h - 8)
        svg = {"w": w, "h": h, "steps": steps, "ki_y": ki_y}

    # 수익률 모의실험 결과 (배치가 저장한 캐시)
    sim = product.sim_result or None

    # 이 상품을 보유 중이면 상품별 낙인 모니터링 (자산별 레벨/버퍼)
    from core.models import KnockInStatus
    ki_statuses = None
    ki_worst_buffer = None
    ki_updated_at = None
    inv = None
    if request.user.is_authenticated:
        inv = (product.investments.filter(user=request.user, status="보유중")
               .prefetch_related("ki_status").first())
    if inv:
        rows = list(inv.ki_status.all())
        if rows:
            for s in rows:
                s.buffer = None if (s.level_pct is None or product.ki is None or product.is_no_ki) \
                    else round(s.level_pct - product.ki, 1)
                if s.updated_at and (ki_updated_at is None or s.updated_at > ki_updated_at):
                    ki_updated_at = s.updated_at
            ki_statuses = sorted(rows, key=lambda s: (s.level_pct if s.level_pct is not None else 999))
            ki_worst_buffer = inv.ki_buffer

    # ── 기초자산 1년 시세 차트 (레벨% = 종가/기준가×100, SVG) ──
    from core import market as _mkt
    chart = None
    assets = _mkt.split_assets(product.assets_raw)
    series_list = []
    # 기준가 산정일 + 거래일 오프셋 (설명서 평가일 우선, 없으면 발행사별 규칙)
    base_date, base_back = _mkt.base_price_date(product)
    for asset in assets[:4]:
        t = _mkt.resolve_ticker(asset)
        if not t:
            continue
        hist = _mkt.fetch_history(t)
        if len(hist) < 10:
            continue
        ref = None
        if base_date:
            past = [c for d, c in hist if d <= base_date]
            ref = past[-1 - base_back] if len(past) > base_back else None
        if ref is None:
            ref = hist[0][1]  # 미발행 상품은 1년 전 시점=100
        hi_d, hi_c = max(hist, key=lambda x: x[1])
        lo_d, lo_c = min(hist, key=lambda x: x[1])
        series_list.append({"asset": asset, "ref": ref,
                            "hi": (hi_d, hi_c), "lo": (lo_d, lo_c),
                            "cur": hist[-1][1],
                            "pts": [(d, c / ref * 100) for d, c in hist]})
    if series_list:
        all_d = [d for s in series_list for d, _ in s["pts"]]
        dmin, dmax = min(all_d), max(all_d)
        span = (dmax - dmin).days or 1
        levels = [v for s in series_list for _, v in s["pts"]]
        marks = [100]
        if product.barrier_first is not None:
            marks.append(product.barrier_first)
        if product.ki is not None and not product.is_no_ki:
            marks.append(product.ki)
        lo = min(min(levels), min(marks)) - 5
        hi = max(max(levels), max(marks)) + 5
        W, H, PL, PR, PT, PB = 720, 260, 44, 10, 14, 30
        pw, ph = W - PL - PR, H - PT - PB

        def _x(d):
            return round(PL + pw * (d - dmin).days / span, 1)

        def _y(v):
            return round(PT + ph * (1 - (v - lo) / (hi - lo)), 1)

        palette = ["#1b64da", "#e8590c", "#0ca678", "#845ef7"]

        def _fmt_price(v):
            return f"{v:,.0f}" if v >= 1000 else f"{v:,.2f}"

        chart_series = []
        for i, s in enumerate(series_list):
            ref = s["ref"]
            (hi_d, hi_c), (lo_d, lo_c) = s["hi"], s["lo"]
            chart_series.append({
                "asset": s["asset"], "color": palette[i % 4],
                "poly": " ".join(f"{_x(d)},{_y(v)}" for d, v in s["pts"]),
                "last": round(s["pts"][-1][1], 1),
                # 실제 가격 정보 (기준가 통화 단위 그대로)
                "ref_price": _fmt_price(ref),
                "cur_price": _fmt_price(s["cur"]),
                "first_price": _fmt_price(ref * product.barrier_first / 100)
                               if product.barrier_first is not None else None,
                "ki_price": _fmt_price(ref * product.ki / 100)
                            if (product.ki is not None and not product.is_no_ki) else None,
                "hi_price": _fmt_price(hi_c), "hi_date": hi_d,
                "hi_x": _x(hi_d), "hi_y": _y(hi_c / ref * 100),
                "lo_price": _fmt_price(lo_c), "lo_date": lo_d,
                "lo_x": _x(lo_d), "lo_y": _y(lo_c / ref * 100),
            })
        chart_lines = [{"label": "기준 100", "y": _y(100), "color": "#868e96", "dash": "4 3"}]
        if product.barrier_first is not None:
            chart_lines.append({"label": f"1차 {product.barrier_first:g}",
                                "y": _y(product.barrier_first), "color": "#e8590c", "dash": "6 3"})
        if product.ki is not None and not product.is_no_ki:
            chart_lines.append({"label": f"낙인 {product.ki:g}",
                                "y": _y(product.ki), "color": "#e03131", "dash": "2 3"})
        # 고점대비 — 발행 시점(기준가 산정일)과 현재 시점을 따로 계산해 둘 다 보여준다.
        # 화면 문구가 "…자리에서 발행"이면 숫자도 발행일 값이어야 한다(2026-08-05 태훈님).
        # 상품 단위 값은 자산별 최댓값 = 가장 덜 빠진 자산 기준(레이더 게이트와 같은 뜻).
        peak = peak_ratios(product)
        # 고점 발행이었나 판정은 템플릿이 issue_max(발행 완료분)로 한다 — v7 게이트가
        # 보는 시점과 맞추기 위함. 미발행분만 now_max로 "이대로면" 이라고 쓴다.
        peak["asset_rows"] = [r for r in peak["rows"] if r["now"] is not None]
        # 발행일(기준가 산정일) 세로 점선 — 차트 구간 안에 있을 때만
        base_x = _x(base_date) if (base_date and dmin <= base_date <= dmax) else None
        chart = {"W": W, "H": H, "series": chart_series, "lines": chart_lines,
                 "based_on_issue": bool(base_date),
                 "range_from": dmin, "range_to": dmax,
                 "base_x": base_x, "top": PT, "bottom": H - PB,
                 "peak": peak}

    return render(request, "core/product_detail.html", {
        "meta_desc": _product_meta_desc(product),
        "product": product, "is_watched": is_watched, "svg": svg,
        "sim": sim, "sim_updated": product.sim_updated,
        "ki_statuses": ki_statuses, "ki_worst_buffer": ki_worst_buffer,
        "ki_updated_at": ki_updated_at,
        "chart": chart,
        "my_inv": inv,  # 보유 중이면 상품 정보에 내 투자금액 표시
        # 유사상품 비교 — 조각 템플릿이 컨텍스트를 통째로 받으므로 여기서 미리
        # 그려 문자열로 넘긴다. 상세 컨텍스트에 키 열두 개를 풀어놓지 않기 위함이다.
        # 같은 조건 상품이 3건 미만이면 None → 카드 자체가 안 나온다.
        "compare_html": _compare_html(request, product),
        "active_nav": "weekly",
    })


# ── 유사상품 비교 ─────────────────────────────────
def product_compare(request, pk):
    """비교 패널 조각만 돌려준다 — 주간 목록의 '비교 N' 모달과 상세 화면이 함께 쓴다.

    ?same=1이면 '같은 기초자산만' 토글이 켜진 상태. 토글을 화면에서 감추는 대신
    서버에 다시 물어보는 이유는 모수가 바뀌면 백분위·중앙값·최저·최고가 전부
    다시 계산돼야 하기 때문이다 — 감추기만 하면 게이지가 거짓말을 한다.

    저장하지 않는다(매주 모수가 바뀐다). 대신 모수 조회는 쿼리 한 번으로 끝낸다.
    """
    product = get_object_or_404(Product, pk=pk)
    ctx = compare_context(product, same_assets=request.GET.get("same") == "1")
    if ctx is None:
        # 버튼이 안 보이는 상품인데 주소로 직접 들어온 경우 — 화면을 깨지 않고 알린다.
        return HttpResponse(
            '<div style="font-size:12.5px;color:var(--text-2);line-height:1.7">'
            '같은 주에 조건이 같은 상품이 충분하지 않아 비교할 수 없습니다.</div>')
    return render(request, "core/_compare_panel.html", ctx)


# ── 프리셋 관리 ───────────────────────────────────
@login_required
def presets(request):
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "delete":
            _scope(Preset.objects.filter(id=request.POST.get("id")), request.user).delete()
            messages.success(request, "프리셋을 삭제했습니다.")
        else:  # create / update
            pid = request.POST.get("id")
            data = dict(
                name=request.POST.get("name", "").strip() or "이름없음",
                issuers=request.POST.getlist("issuers"),
                ki_min=request.POST.get("ki_min") or None,
                ki_max=request.POST.get("ki_max") or None,
                include_no_ki=request.POST.get("include_no_ki") == "on",
                asset_type=request.POST.get("asset_type", "전체"),
                yield_min=request.POST.get("yield_min") or None,
                period_max=request.POST.get("period_max") or None,
                currency=request.POST.get("currency", "전체"),
                notify=request.POST.get("notify") == "on",
            )
            # 텔레그램은 가족 공용 운영 채널이라 일반 회원 프리셋은 태우지 않는다.
            # 가드가 추가(create)에만 있어서, 만든 뒤 수정으로 다시 켤 수 있었다.
            # 화면에서 체크박스를 감추더라도 POST는 직접 올 수 있으므로 서버에서 막는다.
            if not request.user.is_staff:
                data["notify"] = False
            if pid:
                _scope(Preset.objects.filter(id=pid), request.user).update(**data)
                messages.success(request, "프리셋을 수정했습니다.")
            else:
                data["user"] = request.user
                Preset.objects.create(**data)
                messages.success(request, "프리셋을 추가했습니다.")
        return redirect("presets")

    today = date.today()
    active_products = Product.objects.listed().filter(sub_end__gte=today)
    preset_list = []
    for p in _scope(Preset.objects.all(), request.user):
        preset_list.append({"obj": p, "match_count": p.match_queryset(active_products).count()})

    # 발행사 후보 (전체 상품 기준 — 최근 60일)
    issuers = sorted(set(
        Product.objects.listed().filter(sub_end__gte=today - timedelta(days=60))
        .values_list("issuer", flat=True)
    ))

    return render(request, "core/presets.html", {
        "preset_list": preset_list, "issuers": issuers, "active_nav": "presets",
    })


# ── 관심 목록 ─────────────────────────────────────
@login_required
def watchlist(request):
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "clear":  # product_id 없이 옴 → 상단에서 먼저 처리
            n, _ = _scope(WatchItem.objects.all(), request.user).delete()
            messages.success(request, f"관심 목록 {n}건을 초기화했습니다.")
            return redirect("watchlist")
        product = get_object_or_404(Product, pk=request.POST.get("product_id"))
        is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
        if action == "add":
            WatchItem.objects.get_or_create(product=product, user=request.user)
            if not is_ajax:
                messages.success(request, "관심 목록에 등록했습니다.")
        elif action == "remove":
            _scope(WatchItem.objects.filter(product=product), request.user).delete()
            if not is_ajax:
                messages.success(request, "관심 목록에서 해제했습니다.")
        if is_ajax:
            from django.http import JsonResponse
            return JsonResponse({"watched": action == "add"})
        return redirect(request.POST.get("next") or "watchlist")

    items = list(_scope(WatchItem.objects.select_related("product").all(), request.user))
    invested_ids = set(Investment.objects.filter(
        user=request.user, status="보유중").values_list("product_id", flat=True))

    # ── 정렬 (헤더 클릭) ──
    _min_date = date.min
    W_SORT = {
        "issuer": lambda i: i.product.issuer or "",
        "no": lambda i: i.product.product_no or "",
        "assets": lambda i: i.product.assets_raw or "",
        "yield": lambda i: i.product.yield_rate if i.product.yield_rate is not None else -1,
        "ki": lambda i: 999 if i.product.is_no_ki else (i.product.ki if i.product.ki is not None else -1),
        "first": lambda i: i.product.barrier_first if i.product.barrier_first is not None else -1,
        "last": lambda i: i.product.barrier_last if i.product.barrier_last is not None else -1,
        "term": lambda i: i.product.term_months if i.product.term_months is not None else -1,
        "period": lambda i: i.product.period_months if i.product.period_months is not None else -1,
        "loss": lambda i: i.product.loss_prob if i.product.loss_prob is not None else -1,
        "type": lambda i: i.product.asset_type or "",
        "sub_end": lambda i: i.product.sub_end or _min_date,
        "confirm": lambda i: i.product.confirm_date or _min_date,
    }
    # 기본 정렬: 마감 가까운 순 (sub_end 오름차순)
    w_sort = request.GET.get("wsort", "sub_end")
    w_dir = request.GET.get("wdir", "asc")
    if w_sort in W_SORT:
        items.sort(key=W_SORT[w_sort], reverse=(w_dir == "desc"))

    def _wurl(key):
        d = "desc" if (w_sort == key and w_dir == "asc") else "asc"
        return f"?wsort={key}&wdir={d}"

    w_cols = [
        {"key": k, "label": lbl, "num": num, "url": _wurl(k),
         "active": w_sort == k, "dir": w_dir}
        for k, lbl, num in [
            ("issuer", "발행사", False), ("no", "상품번호", False),
            ("assets", "기초자산", False), ("yield", "수익률", True),
            ("ki", "낙인", True), ("first", "1차", True), ("last", "막차", True),
            ("term", "기간", True), ("period", "주기", True),
            ("loss", "손실확률", True), ("type", "유형", False),
            ("sub_end", "마감", True), ("confirm", "숙려마감", True),
        ]
    ]

    return render(request, "core/watchlist.html", {
        "invested_ids": invested_ids,
        "items": items, "w_cols": w_cols, "active_nav": "watchlist",
        "vapid_public_key": settings.VAPID_PUBLIC_KEY,
    })


# ── 포트폴리오 ────────────────────────────────────
# 조작(POST)은 fetch로 처리하고 바뀐 화면 조각만 돌려준다. 전체 새로고침을 하면
# 리다이렉트가 지금 URL의 쿼리스트링(정렬·페이지)을 버려서 보던 화면이 풀린다.
# 다만 JS가 없거나 fetch가 실패하면 예전처럼 일반 폼 제출 + 리다이렉트로 간다.
# 조각의 경계(id="pf-*")는 portfolio.html에 있고, 어느 조각을 갈아끼울지는
# 같은 파일의 PF_SECTIONS(JS)가 정한다 — 서버는 조각을 고르지 않는다.


def _pf_is_fetch(request):
    """fetch가 보낸 요청인지 — 일반 폼 제출과 갈라야 폴백이 성립한다."""
    return request.headers.get("X-Requested-With") == "XMLHttpRequest"


@login_required
def portfolio(request):
    if request.method == "POST":
        _portfolio_action(request)
        if _pf_is_fetch(request):
            return _portfolio_fragment_response(request)
        return redirect(request.POST.get("next") or "portfolio")
    return render(request, "core/portfolio.html", _portfolio_context(request))


def _portfolio_fragment_response(request):
    """페이지 전체 대신 content 블록만 렌더해 JSON으로 준다.

    조각을 액션별로 골라 보내지 않고 항상 전부 보낸다 — 상환 처리 하나가
    통계·리스크·보유·상환완료를 동시에 바꿔서, 액션마다 목록을 관리하면
    빠뜨린 조각이 조용히 낡은 값으로 남는다. 실제로 DOM을 갈아끼우는 건
    내용이 달라진 조각뿐이라(JS에서 비교) 화면 깜빡임은 없다.
    """
    # 메시지는 여기서 비운다 — 안 그러면 세션에 남아 다음 새로고침에 또 뜬다
    storage = messages.get_messages(request)
    texts = [str(m) for m in storage]
    storage.used = True
    ctx = _portfolio_context(request)
    ctx["base_template"] = "core/_pf_fragment_base.html"
    return JsonResponse({
        "html": render_to_string("core/portfolio.html", ctx, request=request),
        "message": " ".join(texts),
    })


def _portfolio_action(request):
    """포트폴리오 화면의 데이터 조작 — 응답 형식과 무관한 부분."""
    action = request.POST.get("action")
    if action == "add":
        product = get_object_or_404(Product, pk=request.POST.get("product_id"))
        Investment.objects.create(
            user=request.user,
            product=product,
            amount=int(request.POST.get("amount", "0").replace(",", "")),
            # 발행일 미입력 시 상품 발행일 기준 — 실현수익 연환산이
            # 발행일~상환일 실경과일로 계산되므로 발행일이 정확한 기준
            invested_at=(request.POST.get("invested_at")
                         or product.issued_on or date.today()),
            broker_account=request.POST.get("broker_account", ""),
            memo=request.POST.get("memo", ""),
        )
        _scope(WatchItem.objects.filter(product=product), request.user).delete()
        messages.success(request, "투자를 등록했습니다.")
    elif action == "redeem":
        inv = get_object_or_404(Investment, pk=request.POST.get("id"), user=request.user)
        inv.status = request.POST.get("status", "조기상환")
        inv.redeemed_at = request.POST.get("redeemed_at") or date.today()
        amt = request.POST.get("redeemed_amount", "").replace(",", "")
        inv.redeemed_amount = int(amt) if amt else None
        inv.save()
        messages.success(request, "상환 처리했습니다.")
    elif action == "edit":
        inv = get_object_or_404(Investment, pk=request.POST.get("id"), user=request.user)
        changed = False
        if "amount" in request.POST:
            a = request.POST.get("amount", "").replace(",", "").strip()
            if a.isdigit() and int(a) > 0:
                inv.amount = int(a)
                changed = True
        if "redeemed_amount" in request.POST:
            r = request.POST.get("redeemed_amount", "").replace(",", "").strip()
            inv.redeemed_amount = int(r) if r.isdigit() else None
            changed = True
        if changed:
            inv.save()
            messages.success(request, "금액을 수정했습니다.")
    elif action in ("dismiss_verdict", "undismiss_verdict"):
        # 판정이 틀렸을 때(증권사 확인 결과 상환이 아닌 경우) 그 회차 재알림만 멈춘다.
        # 회차마다 행이 따로라 다음 회차 판정·알림은 그대로 동작한다.
        inv = get_object_or_404(Investment, pk=request.POST.get("id"), user=request.user)
        v = RedemptionVerdict.objects.filter(
            investment=inv, round_no=request.POST.get("round_no")).first()
        if v:
            on = action == "dismiss_verdict"
            v.dismissed_at = timezone.now() if on else None
            v.save(update_fields=["dismissed_at"])
            messages.success(
                request,
                "판정을 무시했습니다. 이 회차는 다시 알리지 않습니다."
                if on else "판정 무시를 해제했습니다.")
    elif action == "delete":
        Investment.objects.filter(pk=request.POST.get("id"), user=request.user).delete()
        messages.success(request, "투자 기록을 삭제했습니다.")
    elif action == "bulk_delete":
        ids = request.POST.getlist("ids")
        n, _ = Investment.objects.filter(pk__in=ids, user=request.user).delete()
        messages.success(request, f"{n}건의 투자 기록을 삭제했습니다.")


def _portfolio_context(request):
    """화면에 필요한 값 전부 — GET 렌더와 fetch 조각 렌더가 같이 쓴다."""
    invs = (Investment.objects.filter(user=request.user)
            .select_related("product").prefetch_related("ki_status", "verdicts"))
    holding = [i for i in invs if i.status == "보유중"]
    done = [i for i in invs if i.status != "보유중"]
    # 조기상환 실패: 지난 평가에서 배리어 미달이 확정된 건 (통계에는 보유중으로 포함,
    # 리스트 표시만 분리 — 돈은 여전히 들어가 있는 상태이므로)
    missed = [i for i in holding if i.missed_redemption]
    # 상환 확정 대기: 배리어 충족 판정이 났는데 아직 상환 처리가 안 된 건.
    # 보유 목록에 작은 배지로만 있으면 그냥 지나친다 — 2026-08-03 판정된 키움 1863
    # 한 건이 일주일간 방치된 원인이라, 실패 건과 같은 방식으로 맨 위 전용 카드로 뺀다.
    pending = [i for i in holding if i.redemption_pending]
    holding_display = [i for i in holding
                       if not i.missed_redemption and not i.redemption_pending]

    today = date.today()
    month_end = date(today.year, today.month, pycalendar.monthrange(today.year, today.month)[1])
    this_month_evals = 0
    for inv in holding:
        nxt = inv.next_evaluation
        if nxt and today <= nxt["date"] <= month_end:
            this_month_evals += 1

    total_invested = sum(i.amount for i in holding)
    total_redeemed_profit = sum(
        (i.redeemed_amount - i.amount) for i in done if i.redeemed_amount
    )

    # 세전 예상 수익 (보유분이 1차 평가에 전부 조기상환된다고 가정)
    total_expected_pretax = 0
    for i in holding:
        sched = i.schedule
        total_expected_pretax += sched[0]["expected"] if sched else i.amount
    expected_profit_pretax = total_expected_pretax - total_invested
    # 수익률 = 예상 세전수익 ÷ 총 투자금액 (연환산 아님 — 1차 상환까지의 단순 수익률)
    expected_profit_rate = (round(expected_profit_pretax / total_invested * 100, 2)
                            if total_invested else None)

    # 포트폴리오 예상 손실율 = Σ(투자금 × 손실확률) / Σ투자금  (금액 가중평균)
    weighted_loss = 0
    loss_weight = 0  # 손실확률이 있는 투자금 합 (커버리지 표기용)
    for i in holding:
        lp = i.product.loss_prob
        if lp is not None:
            weighted_loss += i.amount * lp
            loss_weight += i.amount
    port_loss_rate = round(weighted_loss / total_invested, 2) if total_invested else None
    loss_coverage_pct = round(loss_weight / total_invested * 100) if total_invested else 0

    # 유형별(종목형/지수형) 보유 분해 — 건수·투자금액
    holding_by_type = portfolio_facts.holding_by_type(holding)

    # ── 리스크 분석 ──────────────────────────────
    risk = portfolio_facts.analyze_risk(holding, total_invested)

    # ── 스트레스 테스트 (계산은 core/portfolio_facts.py — /ask/ 도구와 공용) ──
    stress = portfolio_facts.stress_test(holding, total_invested)

    # ── 낙인 모니터링: 전체 보유 중 위험/경고(버퍼 ≤ 20%p)만 추림 ──
    ki_updated = None
    ki_alerts = []
    for inv in holding:
        worst = inv.worst_ki_status
        for s in inv.ki_status.all():
            if s.updated_at and (ki_updated is None or s.updated_at > ki_updated):
                ki_updated = s.updated_at
        buf = inv.ki_buffer
        if worst is not None and buf is not None and buf <= 20:
            ki_alerts.append({"inv": inv, "worst": worst, "buffer": buf})
    ki_alerts.sort(key=lambda a: a["buffer"])  # 위험한 순
    has_ki_data = ki_updated is not None

    # 투자 등록 폼용 상품 후보 (최근 청약 상품)
    # 투자등록 후보 = 관심목록에 담아둔 상품만 (가족=공용+본인, 회원=본인)
    # .listed()를 걸지 않는다 — 본인이 직접 담은 관심목록이 원본이라, 여기서
    # 거르면 담아둔 상품이 등록 목록에서만 사라져 이유를 알 수 없게 된다.
    candidates = Product.objects.filter(
        id__in=_scope(WatchItem.objects.all(), request.user).values_list("product_id", flat=True)
    ).order_by("-sub_end", "issuer")

    # ── 보유 리스트 정렬 ──
    def _pretax(i):
        s = i.schedule
        return s[0]["expected"] if s else 0

    H_SORT = {
        "issuer": lambda i: (i.product.issuer or ""),
        "assets": lambda i: (i.product.assets_raw or ""),
        "amount": lambda i: i.amount or 0,
        "yield": lambda i: i.product.yield_rate if i.product.yield_rate is not None else -1,
        "issue": lambda i: (i.product.issued_on or date.max),
        "ki": lambda i: (999 if i.product.is_no_ki or i.product.ki is None else i.product.ki),
        # Product에는 stepdown_barriers가 없다(HistoricalIssue 필드) — 예전엔
        # AttributeError로 이 정렬이 500을 냈다. 1차 배리어는 barrier_first다.
        "b1": lambda i: (i.product.barrier_first
                         if i.product.barrier_first is not None else 999),
        "next": lambda i: (i.next_evaluation["date"] if i.next_evaluation else date.max),
        "pretax": _pretax,
        "loss": lambda i: (i.product.loss_prob if i.product.loss_prob is not None else -1),
    }
    h_sort = request.GET.get("hsort", "next")
    if h_sort not in H_SORT:
        h_sort = "next"
    h_dir = request.GET.get("hdir", "asc")
    holding_display.sort(key=H_SORT[h_sort], reverse=(h_dir == "desc"))
    missed.sort(key=lambda i: i.missed_redemption.eval_date)  # 오래 놓친 순
    # 무시한 건은 아래로, 나머지는 오래 방치된 순
    pending.sort(key=lambda i: (i.redemption_pending.dismissed_at is not None,
                                i.redemption_pending.eval_date))

    # ── 상환 완료 리스트 정렬 ──
    # 상환금액·실현수익률·상환일은 비어 있을 수 있다(상태만 바꾸고 상환금을 넣지
    # 않은 과거 기록). None끼리 비교하면 TypeError로 500이 나므로 H_SORT처럼
    # 실제 값보다 항상 작은/큰 대체값을 넣는다. 실현수익률은 음수(-40% 등)가
    # 정상 값이라 -1을 못 쓴다 — 그래서 -inf다.
    _MISSING = float("-inf")
    D_SORT = {
        "issuer": lambda i: (i.product.issuer or ""),
        "amount": lambda i: i.amount or 0,
        "redeemed": lambda i: (i.redeemed_amount
                               if i.redeemed_amount is not None else _MISSING),
        "realized": lambda i: (i.realized_return_pct
                               if i.realized_return_pct is not None else _MISSING),
        "redeemed_at": lambda i: (i.redeemed_at or date.max),
    }
    # 정렬 파라미터가 없거나(또는 모르는 컬럼이면) 상환일 내림차순(최근 상환
    # 먼저)이 기본이다 — 조 팀장 지시(2026-08-11). 컬럼을 명시적으로 골랐는데
    # 방향(ddir)만 안 왔으면 그 컬럼의 기본인 오름차순을 쓴다 — 페이지 기본값
    # "desc"를 아무 컬럼에나 덮어씌우면 방금 누른 컬럼 방향이 뒤집혀 버린다.
    _raw_dsort = request.GET.get("dsort", "")
    if _raw_dsort in D_SORT:
        d_sort = _raw_dsort
        d_dir = request.GET.get("ddir", "asc")
    else:
        d_sort = "redeemed_at"
        d_dir = request.GET.get("ddir", "desc")
    done.sort(key=D_SORT[d_sort], reverse=(d_dir == "desc"))

    def _hsort_url(key):
        d = "desc" if (h_sort == key and h_dir == "asc") else "asc"
        # 상환완료 정렬을 같이 실어 보낸다 — 위 표를 정렬했다고 아래 표 정렬이
        # 풀리면 안 되기 때문이다. d_sort는 기본값(상환일 내림차순)도 포함해
        # 항상 값이 있으므로 매번 실어 보낸다.
        return f"?hsort={key}&hdir={d}&dsort={d_sort}&ddir={d_dir}&psize={page_size}"

    def _dsort_url(key):
        d = "desc" if (d_sort == key and d_dir == "asc") else "asc"
        return f"?dsort={key}&ddir={d}&hsort={h_sort}&hdir={h_dir}&psize={page_size}"

    # ── 페이지네이션 ──
    from django.core.paginator import Paginator
    try:
        page_size = int(request.GET.get("psize", 10))
    except (ValueError, TypeError):
        page_size = 10
    if page_size not in (10, 20, 50, 100):
        page_size = 10

    h_page = Paginator(holding_display, page_size).get_page(request.GET.get("hpage"))
    d_page = Paginator(done, page_size).get_page(request.GET.get("dpage"))

    h_cols = [
        {"key": k, "label": lbl, "num": num, "url": _hsort_url(k),
         "active": h_sort == k, "dir": h_dir}
        for k, lbl, num in [
            ("issuer", "상품", False), ("assets", "기초자산", False),
            ("amount", "투자금액", True), ("yield", "수익률", True),
            ("ki", "낙인", True), ("b1", "1차 조기상환", True),
            ("issue", "발행일", False),
            ("next", "다음 평가일", False), ("pretax", "예상상환금", True),
            ("loss", "손실확률", True),
        ]
    ]

    # key가 None인 칸(상태)은 정렬 링크 없이 글자만 — 값이 네 종류뿐이라 실익이 없다
    d_cols = [
        {"key": k, "label": lbl, "num": num, "url": _dsort_url(k) if k else "",
         "active": bool(k) and d_sort == k, "dir": d_dir}
        for k, lbl, num in [
            ("issuer", "상품", False), (None, "상태", False),
            ("amount", "투자금액", True), ("redeemed", "상환금액", True),
            ("realized", "실현수익률", True), ("redeemed_at", "상환일", False),
        ]
    ]

    # ── 실현수익 실적 (상환완료 + 상환금 입력된 건) ──
    realized = [i for i in done if i.redeemed_amount is not None and i.redeemed_at]
    perf = None
    if realized:
        from collections import defaultdict
        total_in = sum(i.amount for i in realized)
        total_out = sum(i.redeemed_amount for i in realized)
        profit = total_out - total_in
        # 연환산: 각 건의 보유일수 가중 수익률 평균
        ann_rates = []
        for i in realized:
            days = (i.redeemed_at - i.invested_at).days if i.invested_at else None
            if days and days > 0:
                ann_rates.append((i.redeemed_amount - i.amount) / i.amount * 365 / days * 100)
        monthly = defaultdict(int)
        for i in realized:
            monthly[i.redeemed_at.strftime("%Y.%m")] += i.redeemed_amount - i.amount
        months = sorted(monthly)[-12:]
        max_abs = max(abs(monthly[m]) for m in months) or 1
        bars = [{"label": m, "value": monthly[m],
                 "h": round(abs(monthly[m]) / max_abs * 100, 1),
                 "neg": monthly[m] < 0} for m in months]
        perf = {
            "count": len(realized),
            "profit": profit,
            "rate": round(profit / total_in * 100, 2) if total_in else 0,
            "ann_rate": round(sum(ann_rates) / len(ann_rates), 2) if ann_rates else None,
            "bars": bars,
        }

    return {
        "base_template": "core/base.html",
        "vapid_public_key": settings.VAPID_PUBLIC_KEY,
        "perf": perf,
        "h_page": h_page, "d_page": d_page,
        "holding_count": len(holding_display), "done_count": len(done),
        "missed": missed, "missed_count": len(missed),
        # 무시한 건은 '대기' 건수에서 뺀다 — 할 일 개수를 부풀리지 않기 위해서다
        "pending": pending,
        "pending_count": sum(1 for i in pending
                             if i.redemption_pending.dismissed_at is None),
        "dismissed_count": sum(1 for i in pending
                               if i.redemption_pending.dismissed_at is not None),
        "holding_total_count": len(holding),  # 빈 상태 문구용 (보유 전체)
        "h_cols": h_cols, "d_cols": d_cols, "page_size": page_size,
        "total_invested": total_invested,
        "holding_by_type": holding_by_type,
        "this_month_evals": this_month_evals,
        "total_redeemed_profit": total_redeemed_profit,
        "total_expected_pretax": total_expected_pretax,
        "expected_profit_pretax": expected_profit_pretax,
        "expected_profit_rate": expected_profit_rate,
        "port_loss_rate": port_loss_rate,
        "loss_coverage_pct": loss_coverage_pct,
        "risk": risk,
        "ki_updated": ki_updated,
        "has_ki_data": has_ki_data,
        "ki_alerts": ki_alerts,
        "stress": stress,
        "candidates": candidates,
        "today": today,
        "active_nav": "portfolio",
    }


# ── 포트폴리오 엑셀 양식 다운로드 ─────────────────
PORTFOLIO_COLS = ["발행사", "상품번호", "투자금액(원)", "발행일(YYYY-MM-DD)", "증권사/계좌", "메모"]


@login_required
def portfolio_template(request):
    """투자내역 일괄등록 엑셀 양식 다운로드."""
    import io
    import openpyxl
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "투자내역"
    ws.append(PORTFOLIO_COLS)
    # 예시 행 (안내용, 업로드 시 발행사가 실제 매칭 안되면 자동 무시됨)
    ws.append(["키움증권", "1965", 10000000, "2026-07-16", "키움 CMA", "예시 행 — 삭제 후 작성"])
    widths = [14, 10, 14, 18, 14, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    guide = wb.create_sheet("작성안내")
    for row in [
        ["ELS 레이더 — 투자내역 일괄등록 양식"],
        [""],
        ["1. '투자내역' 시트에 한 행씩 입력하세요."],
        ["2. 발행사 + 상품번호로 수집된 상품과 자동 매칭합니다."],
        ["   (주간청약/상품 목록에 있는 발행사·상품번호와 동일하게 입력)"],
        ["3. 투자금액은 숫자만 (원 단위). 예: 10000000"],
        ["4. 발행일은 YYYY-MM-DD. 비우면 오늘 날짜로 등록됩니다."],
        ["5. 증권사/계좌·메모는 선택입니다."],
        ["6. 예시 행은 삭제하고 업로드하세요."],
        [""],
        ["※ 매칭 실패한 행은 등록되지 않고 결과에 표시됩니다."],
    ]:
        guide.append(row)
    guide.column_dimensions["A"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="ELS_투자내역_양식.xlsx"'
    return resp


@login_required
def watchlist_export(request):
    """관심목록을 xlsx로 다운로드 (본인 범위)."""
    import io

    import openpyxl
    from django.http import HttpResponse

    cols = ["발행사", "상품번호", "기초자산", "수익률(%)", "낙인", "1차", "막차",
            "기간", "주기", "손실확률(%)", "유형", "청약마감", "숙려마감", "메모", "보유"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "관심목록"
    ws.append(cols)

    items = _scope(WatchItem.objects.select_related("product").all(), request.user)
    invested_ids = set(Investment.objects.filter(
        user=request.user, status="보유중").values_list("product_id", flat=True))
    for item in items:
        p = item.product
        ws.append([
            p.issuer, p.product_no, p.assets_raw or "",
            p.yield_rate if p.yield_rate is not None else "",
            p.ki_display,
            p.barrier_first if p.barrier_first is not None else "",
            p.barrier_last if p.barrier_last is not None else "",
            p.term_display,
            p.period_display,
            p.loss_prob if p.loss_prob is not None else "",
            p.asset_type or (p.structure_label or ""),
            p.sub_end.strftime("%Y-%m-%d") if p.sub_end else "",
            p.confirm_date.strftime("%Y-%m-%d") if p.confirm_date else "",
            item.memo or "",
            "보유중" if p.id in invested_ids else "",
        ])
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(10, len(c) + 4)

    buf = io.BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="watchlist_{date.today():%Y%m%d}.xlsx"'
    return resp


@login_required
def portfolio_export(request):
    """보유·상환 내역을 조 팀장 관리 양식(17열)으로 다운로드 — **본인 데이터만**.

    첫 시트 '보유계약'이 조 팀장이 쓰던 엑셀 양식이고, 둘째 시트 '상세'는 종전
    다운로드 형식(손실확률·다음평가일 등)을 그대로 남겨 둔 것이다.
    """
    import io

    import openpyxl
    from django.http import HttpResponse

    from core import portfolio_export as pfx

    mine = (Investment.objects.filter(user=request.user)
            .select_related("product"))

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "보유계약"
    sheet.append(pfx.COLUMNS)
    for row in pfx.build_rows(mine):
        sheet.append(row)
    for i, w in enumerate(pfx.COLUMN_WIDTHS, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    sheet.freeze_panes = "A2"

    cols = ["발행사", "상품번호", "기초자산", "투자금액(원)", "수익률(%)", "낙인",
            "주기(개월)", "1차까지(개월)", "다음평가일", "예상상환금", "손실확률(%)",
            "발행일", "만기일", "스케줄"]

    ws = wb.create_sheet("상세")
    ws.append(cols)

    holding = sorted(
        [i for i in mine if i.status == "보유중"],
        key=lambda i: (i.next_evaluation["date"] if i.next_evaluation else date.max),
    )
    for inv in holding:
        p = inv.product
        nxt = inv.next_evaluation
        badge = inv.schedule_badge or "확정"
        ws.append([
            p.issuer, p.product_no, p.assets_raw or "",
            inv.amount,
            p.yield_rate if p.yield_rate is not None else "",
            p.ki_display,
            p.period_months if p.period_months is not None else "",
            p.first_eval_months if p.first_eval_months is not None else "",
            (nxt["date"].strftime("%Y-%m-%d") if nxt else ""),
            (nxt["expected"] if nxt and nxt["expected"] else ""),
            p.loss_prob if p.loss_prob is not None else "",
            (p.issued_on.strftime("%Y-%m-%d") if p.issued_on else ""),
            (p.expiry_date.strftime("%Y-%m-%d") if p.expiry_date else ""),
            badge,
        ])

    widths = [12, 10, 22, 14, 9, 6, 9, 11, 12, 14, 9, 12, 12, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    fname = f"ELS_보유계약_{date.today():%Y%m%d}.xlsx"
    from urllib.parse import quote
    resp["Content-Disposition"] = (
        f"attachment; filename=\"portfolio.xlsx\"; "
        f"filename*=UTF-8''{quote(fname)}"
    )
    return resp


def portfolio_sync(request):
    """구글 시트가 하루 한 번 읽어 가는 동기화 피드 — 토큰 인증, 로그인 불필요.

    왜 로그인이 아니라 토큰인가
        Google Apps Script는 우리 세션 쿠키를 들고 다니지 못한다. 서버가 시트에
        쓰는 반대 방향을 택하면 구글 서비스 계정 키를 EC2에 두어야 하는데, 그
        키는 시트 하나가 아니라 계정이 닿는 문서 전부의 열쇠라 위험이 훨씬 크다.
        그래서 '읽기 전용 피드 + 토큰' 쪽으로 갔다.

    토큰을 못 맞히면 전부 404다(403이 아니라). 403은 '주소는 맞다'는 확인을
    공짜로 주는 셈이라, 무작위 탐색에 힌트를 남기지 않으려고 없는 주소처럼
    군다. 대신 서버 로그에 이유를 남긴다 — 조 팀장이 진단할 곳은 로그다.

    토큰은 헤더(X-Sync-Token)로 받는 것이 정석이다. 쿼리스트링(?token=)도
    받지만 이쪽은 nginx 접근 로그에 토큰이 그대로 남으므로 브라우저로 한 번
    확인해 볼 때만 쓰고, 스크립트는 헤더를 쓴다.
    """
    from core import portfolio_sync as pfs

    expected = (getattr(settings, "SHEET_SYNC_TOKEN", "") or "").strip()
    if not expected:
        logger.warning("시트 동기화 요청을 받았으나 SHEET_SYNC_TOKEN이 비어 있다 — 404로 응답")
        raise Http404

    given = (request.headers.get("X-Sync-Token")
             or request.GET.get("token") or "").strip()
    # compare_digest: 문자열 비교 시간으로 토큰을 한 글자씩 알아내는 걸 막는다
    if not given or not hmac.compare_digest(given, expected):
        logger.warning("시트 동기화 토큰 불일치 — 404로 응답 (제시된 길이 %d)", len(given))
        raise Http404

    username = (getattr(settings, "SHEET_SYNC_USERNAME", "") or "admin").strip()
    mine = (Investment.objects.filter(user__username=username)
            .select_related("product"))
    # 보유중·상환완료를 모두 낸다. 시트에 '완료' 행이 그대로 남아 있어서,
    # 보유중만 보내면 상환된 행의 상태가 영영 갱신되지 않는다.
    payload = pfs.build_payload(mine, username=username)
    if not payload["count"]:
        logger.warning("시트 동기화 피드가 비어 있다 — SHEET_SYNC_USERNAME=%r 확인 필요", username)

    resp = JsonResponse(payload, json_dumps_params={"ensure_ascii": False})
    resp["Cache-Control"] = "no-store"
    resp["X-Robots-Tag"] = "noindex, nofollow"
    return resp


def _match_product_for_investment(issuer, product_no):
    """발행사+상품번호로 Product 매칭.

    중복 행(같은 issuer·product_no, sub_end만 다름)이 있을 때 배리어(스케줄 정보)가
    있는 '정상 행'을 우선 선택한다. 그다음 최신 sub_end. 이렇게 해야 스케줄이 빈
    껍데기 행에 투자가 연결되는 문제(예: 미래에셋 37858)를 막는다.
    """
    qs = Product.objects.filter(
        issuer=str(issuer).strip(), product_no=str(product_no).strip()
    )
    candidates = list(qs)
    if not candidates:
        return None

    def _sort_key(p):
        has_barriers = 1 if (p.barriers_raw and len(p.barriers_raw) > 0) else 0
        sub = p.sub_end or date.min  # None은 가장 뒤로
        return (has_barriers, sub, p.id)

    return max(candidates, key=_sort_key)


@login_required
def portfolio_upload(request):
    """엑셀로 투자내역 일괄 등록."""
    if request.method != "POST":
        return redirect("portfolio")

    import openpyxl

    f = request.FILES.get("excel")
    if not f or not f.name.lower().endswith((".xlsx", ".xlsm")):
        messages.error(request, "엑셀 파일(.xlsx)을 선택해주세요.")
        return redirect("portfolio")

    try:
        wb = openpyxl.load_workbook(f, data_only=True)
    except Exception:  # noqa: BLE001
        # 예외 원문(openpyxl·zipfile의 영문 메시지, 서버 파일 경로가 섞여 나온다)이
        # 회원 화면에 그대로 찍히던 것을 로그로 돌린다. 화면에는 사람 말 한 줄만.
        # (2026-08-06)
        logger.exception("포트폴리오 엑셀 업로드 실패: %s", getattr(f, "name", ""))
        messages.error(request, "파일을 읽을 수 없습니다. 형식을 확인해 주세요.")
        return redirect("portfolio")

    ws = wb["투자내역"] if "투자내역" in wb.sheetnames else wb.worksheets[0]

    created = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(c is None for c in row):
            continue
        issuer = row[0]
        product_no = row[1] if len(row) > 1 else None
        amount = row[2] if len(row) > 2 else None
        invested = row[3] if len(row) > 3 else None
        broker = row[4] if len(row) > 4 else ""
        memo = row[5] if len(row) > 5 else ""

        if not issuer or product_no is None or amount in (None, ""):
            errors.append(f"{i}행: 발행사·상품번호·투자금액은 필수입니다.")
            continue

        product = _match_product_for_investment(issuer, product_no)
        if not product:
            errors.append(
                f"{i}행: 발행사 '{str(issuer).strip()}' + 상품번호 '{str(product_no).strip()}' "
                f"에 해당하는 수집 상품이 없어 등록하지 못했습니다 "
                f"(발행사명·상품번호를 목록과 동일하게 입력했는지 확인)."
            )
            continue

        try:
            amount_int = int(str(amount).replace(",", "").replace("원", "").strip())
        except (ValueError, TypeError):
            errors.append(f"{i}행: 투자금액 '{amount}'을 숫자로 읽을 수 없습니다.")
            continue

        # 날짜 미기재 시 상품 발행일로 — 오늘 날짜 폴백은 발행일 표시·연환산을 오염시킴
        inv_date = _parse_invest_date(invested) or product.issued_on or date.today()

        Investment.objects.create(
            user=request.user, product=product, amount=amount_int,
            invested_at=inv_date, broker_account=str(broker or "")[:100],
            memo=str(memo or "")[:200],
        )
        WatchItem.objects.filter(product=product).delete()
        created += 1

    if created:
        messages.success(request, f"{created}건의 투자를 등록했습니다.")
    if errors:
        messages.error(request, "일부 행을 건너뛰었습니다: " + " / ".join(errors[:8])
                        + (f" 외 {len(errors)-8}건" if len(errors) > 8 else ""))
    if not created and not errors:
        messages.error(request, "등록할 데이터가 없습니다.")
    return redirect("portfolio")


def _parse_invest_date(val):
    """엑셀 셀값 → date. datetime/문자열/None 처리."""
    from datetime import datetime as _dt
    if val is None or val == "":
        return None
    if hasattr(val, "date"):  # datetime
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip().replace(".", "-").replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return _dt.strptime(s[:10], "%Y-%m-%d").date()
        except ValueError:
            pass
    return None


# ── 시장 트렌드 ───────────────────────────────────
def market_trend(request):
    """주차별 평균 수익률·KI 추이 (sub_end 기준, 최근 20주)."""
    from collections import defaultdict

    weeks_n = 20
    qs = Product.objects.listed().filter(sub_end__isnull=False)
    buckets = defaultdict(list)
    for p in qs:
        monday = p.sub_end - timedelta(days=p.sub_end.weekday())
        buckets[monday].append(p)

    ordered = sorted(buckets)[-weeks_n:]
    rows = []
    for wk in ordered:
        ps = buckets[wk]
        ys = [p.yield_rate for p in ps if p.yield_rate is not None]
        kis = [p.ki for p in ps if p.ki is not None and not p.is_no_ki]
        rows.append({
            "week": wk,
            "count": len(ps),
            "avg_yield": round(sum(ys) / len(ys), 1) if ys else None,
            "avg_ki": round(sum(kis) / len(kis), 1) if kis else None,
        })

    # ── SVG 좌표 계산 ──
    W, H = 720, 240
    PAD_L, PAD_R, PAD_T, PAD_B = 44, 44, 20, 40
    plot_w = W - PAD_L - PAD_R
    plot_h = H - PAD_T - PAD_B

    def _line(key, vmin, vmax):
        vals = [r[key] for r in rows if r[key] is not None]
        if not vals:
            return [], vmin, vmax
        lo = vmin if vmin is not None else min(vals)
        hi = vmax if vmax is not None else max(vals)
        span = (hi - lo) or 1
        pts = []
        n = len(rows)
        for i, r in enumerate(rows):
            if r[key] is None:
                continue
            x = PAD_L + (plot_w * i / max(n - 1, 1))
            y = PAD_T + plot_h * (1 - (r[key] - lo) / span)
            pts.append({"x": round(x, 1), "y": round(y, 1), "v": r[key],
                        "week": r["week"], "count": r["count"]})
        return pts, lo, hi

    yield_pts, y_lo, y_hi = _line("avg_yield", None, None)
    ki_pts, k_lo, k_hi = _line("avg_ki", None, None)

    def _polyline(pts):
        return " ".join(f"{p['x']},{p['y']}" for p in pts)

    # 추세 요약 (첫→마지막)
    trend = None
    if len(yield_pts) >= 2:
        diff = yield_pts[-1]["v"] - yield_pts[0]["v"]
        trend = {
            "yield_diff": round(diff, 1),
            "yield_up": diff >= 0,
            "ki_diff": round(ki_pts[-1]["v"] - ki_pts[0]["v"], 1) if len(ki_pts) >= 2 else None,
        }

    # ── 레이더 신호 성과 검증 (verify_radar가 채운 RadarVerdict 집계) ──
    from core.models import PRINCIPAL_PROTECTED_TYPES, RadarVerdict
    BADGE_TIERS = ("타겟 신호", "아주 강한 신호", "강한 신호")
    # 대조군('배지 없음')에서도 원금지급형을 뺀다 — 배지가 붙을 수 없는 상품이
    # 대조군에만 쌓이면 비교가 기울고, 화면에서 뺀 상품의 성적이 화면에 남는다.
    # 이미 쌓인 행은 verify_radar 배치가 계속 만들지만 여기서 걸러 낸다.
    # (2026-08-06 실측: ELB 판정확정 24건 · 대조군 적중률 83.1% → 83.8%)
    verdicts = list(RadarVerdict.objects.select_related("product")
                    .exclude(product__product_type__in=PRINCIPAL_PROTECTED_TYPES))
    badges = [v for v in verdicts if v.tier in BADGE_TIERS]

    radar = None
    if verdicts:
        evaluated = [v for v in badges if v.met is not None]
        hit = [v for v in evaluated if v.met]
        radar_stats = {
            "total": len(badges),
            "evaluated": len(evaluated),
            "hit": len(hit),
            "hit_rate": round(len(hit) / len(evaluated) * 100, 1) if evaluated else None,
        }

        # 주차별(최근 12주) 배지 상품 적중/미충족/대기 — 막대 높이는 상품 수 비례
        wk_map = defaultdict(lambda: {"hit": 0, "miss": 0, "wait": 0})
        for v in badges:
            b = wk_map[v.week_monday]
            if v.met is True:
                b["hit"] += 1
            elif v.met is False:
                b["miss"] += 1
            else:
                b["wait"] += 1
        # 전체 주차 표시 — 판정 확정이 몰린 초기(1~5월) 주차가 최근 12주 창에
        # 잘려 전부 '대기'만 보이던 문제
        recent_weeks = sorted(wk_map)
        max_total = max((sum(wk_map[w].values()) for w in recent_weeks), default=0) or 1
        BAR_MAX = 96
        radar_weeks = []
        for w in recent_weeks:
            b = wk_map[w]
            total = b["hit"] + b["miss"] + b["wait"]
            done = b["hit"] + b["miss"]
            radar_weeks.append({
                "week": w, "total": total,
                "hit": b["hit"], "miss": b["miss"], "wait": b["wait"],
                "hit_h": round(b["hit"] / max_total * BAR_MAX),
                "miss_h": round(b["miss"] / max_total * BAR_MAX),
                "wait_h": round(b["wait"] / max_total * BAR_MAX),
                "rate": round(b["hit"] / done * 100) if done else None,
            })

        # 등급별 1차 적중률 (배지 2등급 + 대조군 '없음')
        grade_defs = [
            ("타겟 신호", "타겟 신호 (v7)", "#1B64DA"),
            ("아주 강한 신호", "아주 강한 신호 (v6)", "#1B64DA"),
            ("강한 신호", "강한 신호 (v6)", "#3182F6"),
            ("없음", "배지 없음 (대조군)", "#8B95A1"),
        ]
        radar_grades = []
        for tier, label, color in grade_defs:
            ev = [v for v in verdicts if v.tier == tier and v.met is not None]
            ht = [v for v in ev if v.met]
            radar_grades.append({
                "label": label, "color": color,
                "total": len(ev), "hit": len(ht),
                "rate": round(len(ht) / len(ev) * 100, 1) if ev else None,
            })

        # 최근 판정 내역 10건 — 판정 확정건만 (평가 전 대기는 미래 평가일이라 제외)
        radar_recent = sorted(
            (v for v in badges if v.met is not None),
            key=lambda v: v.eval_date, reverse=True)[:10]

        radar = {
            "stats": radar_stats,
            "weeks": radar_weeks,
            "grades": radar_grades,
            "recent": radar_recent,
        }

    return render(request, "core/trend.html", {
        "meta_desc": ("최근 20주간 청약된 ELS의 평균 연 수익률과 낙인 배리어가 어떻게 움직였는지, "
                      "레이더 신호의 실제 조기상환 성적과 함께 봅니다."),
        "rows": rows,
        "yield_pts": yield_pts, "yield_poly": _polyline(yield_pts),
        "ki_pts": ki_pts, "ki_poly": _polyline(ki_pts),
        "y_lo": y_lo, "y_hi": y_hi, "k_lo": k_lo, "k_hi": k_hi,
        "W": W, "H": H, "trend": trend,
        "radar": radar,
        "active_nav": "trend",
    })


# ── 기초자산별 공개 페이지 (로그인 불필요, 검색 유입용) ──────────────
def asset_list(request):
    """기초자산 허브 목록 — 상위 10개 카드."""
    return render(request, "core/asset_list.html", {
        "meta_desc": "삼성전자·SK하이닉스·KOSPI200 등 기초자산별로 최근 발행된 ELS 조건을 모아봅니다.",
        "assets": TOP_ASSETS,
    })


def asset_detail(request, slug):
    """기초자산별 발행 현황 — 로그인 없이 본다."""
    ctx = asset_context(slug)
    if ctx is None:
        raise Http404
    ctx["meta_desc"] = (
        f"{ctx['name']}를 기초자산으로 하는 ELS 발행 현황입니다. "
        f"최근 {ctx['window_days']}일 {ctx['count']}건, 낙인·쿠폰 조건과 발행사 구성을 정리했습니다."
    )
    return render(request, "core/asset_detail.html", ctx)


# ── 상환 캘린더 ───────────────────────────────────
# 조회 가능한 연도. date가 다루는 1~9999에서 앞뒤 한 해씩 물러선다 —
# 이전/다음 달 링크가 year±1을 만들기 때문이다.
MIN_YEAR, MAX_YEAR = date.min.year + 1, date.max.year - 1


@login_required
def redemption_calendar(request):
    today = date.today()
    # 주간 청약의 ?w와 같은 자리 — GET 문자열을 그대로 int()에 넣고 있었다.
    # 2026-08-07 실측: ?y=abc·?m=abc는 ValueError, ?m=13·?m=0은 달력 생성에서
    # IllegalMonthError, ?y=0·?y=999999는 date() 범위 초과로 전부 500이었다.
    # 무효하면 조용히 이번 달로 돌아온다.
    year = _int_param(request.GET.get("y"), today.year, lo=MIN_YEAR, hi=MAX_YEAR)
    month = _int_param(request.GET.get("m"), today.month, lo=1, hi=12)

    prev_y, prev_m = (year - 1, 12) if month == 1 else (year, month - 1)
    next_y, next_m = (year + 1, 1) if month == 12 else (year, month + 1)

    # 이 달의 평가 이벤트 수집
    # 보유중만 가져오던 자리다 — 상환을 확정하는 순간 status가 바뀌어 달력에서
    # 통째로 사라졌다(2026-08-18 실측: 확정한 4건이 그날로 증발). 끝난 건도 남기되
    # 상환된 회차까지만 그린다. 그 이후 회차는 실제로 오지 않기 때문이다.
    events = {}  # day -> [event]
    invs = (Investment.objects.filter(user=request.user)
            .select_related("product").prefetch_related("verdicts"))
    for inv in invs:
        last_n = inv.redeemed_round  # 보유중이면 None = 전 회차 그대로
        for row in inv.schedule:
            if last_n is not None and row["n"] > last_n:
                continue  # 상환 뒤 회차 — 오지 않을 평가일이라 그리지 않는다
            d = row["date"]
            if d.year == year and d.month == month:
                events.setdefault(d.day, []).append({
                    "inv": inv, "n": row["n"],
                    "barrier": row["barrier"], "expected": row["expected"],
                    "badge": inv.schedule_badge,
                    "is_past": d < today,
                    # 상환이 끝난 투자의 이벤트 — 요약 집계에서 뺀다
                    "done": last_n is not None,
                    # 상환이 일어난 그 회차에만 상태를 그대로 붙인다.
                    # 앞 회차는 상환되지 않은 지난 평가라 '지난'이 맞다.
                    "done_label": inv.status if row["n"] == last_n else "",
                })

    cal = pycalendar.Calendar(firstweekday=0)  # 월요일 시작
    weeks = []
    for week in cal.monthdayscalendar(year, month):
        row = []
        for day in week:
            row.append({
                "day": day or "",
                "is_today": bool(day) and date(year, month, day) == today,
                "is_past": bool(day) and date(year, month, day) < today,
                "events": events.get(day, []) if day else [],
            })
        weeks.append(row)

    # ── 이 달 결과 요약 ──
    # 한 투자가 같은 달에 두 번 평가받으면 첫 회차에서 상환되고 끝나므로
    # 대표 1건만 집계한다(투자금액·상환금 중복 방지). 대표는 가장 이른 회차 —
    # 다만 상환이 일어난 회차가 같은 달에 있으면 그 회차를 대표로 올린다.
    # 실제 결과가 나온 자리가 거기이기 때문이다.
    first_ev = {}
    for day in sorted(events):
        for ev in events[day]:
            prev = first_ev.get(ev["inv"].id)
            if prev is None or (ev["done_label"] and not prev[1]["done_label"]):
                first_ev[ev["inv"].id] = (day, ev)
    summary = None
    if first_ev:
        rows = [ev for _, ev in first_ev.values()]
        invested = sum(ev["inv"].amount for ev in rows)
        by_type = {"종목형": {"amount": 0, "count": 0}, "지수형": {"amount": 0, "count": 0}}
        # 상환이 끝난 건도 집계에 넣는다(2026-08-19 조 팀장 지시). 하루 전까지는
        # '앞으로의 예측'이라며 뺐지만, 끝난 건은 결과를 이미 알고 있으므로
        # 추정치가 아니라 실제 상환금(redeemed_amount)을 쓴다.
        # 숫자를 지어내지 않는다 — 다음 둘은 건수·투자금액에만 남기고
        # 금액 집계(상환금·수익·수익률)에서는 뺀다:
        #   · 상환은 끝났는데 실제 상환금이 등록되지 않은 건
        #   · 상환 없이 지나간 회차 (뒤 회차에서 상환된 건의 앞 회차)
        real_sum = est_sum = 0
        real_n = est_n = 0
        base = 0  # 상환금을 산출한 건들의 투자금액 — 세전수익·수익률의 분모
        for ev in rows:
            inv = ev["inv"]
            b = by_type.get(inv.product.asset_type)
            if b:
                b["amount"] += inv.amount
                b["count"] += 1
            if ev["done_label"]:
                if inv.redeemed_amount is not None:
                    real_sum += inv.redeemed_amount
                    real_n += 1
                    base += inv.amount
            elif not ev["done"] and ev["expected"]:
                est_sum += ev["expected"]
                est_n += 1
                base += inv.amount
        profit = real_sum + est_sum - base
        summary = {
            "count": len(rows),
            # 세 갈래는 서로 겹치지 않는다: 상환 완료 / 지난 평가 / 예정
            "done_n": sum(1 for ev in rows if ev["done_label"]),
            "past": sum(1 for ev in rows if ev["is_past"] and not ev["done_label"]),
            "upcoming": sum(1 for ev in rows if not ev["is_past"] and not ev["done_label"]),
            "invested": invested,
            "redeem_total": real_sum + est_sum,
            "real_n": real_n, "real_sum": real_sum,
            "est_n": est_n, "est_sum": est_sum,
            "amount_n": real_n + est_n,
            # 상환금을 알 수 없어 금액 집계에서 뺀 건수 (건수·투자금액에는 남아 있다)
            "no_amount_n": len(rows) - (real_n + est_n),
            "profit": profit,
            "by_type": by_type,
            # 수익률 = 세전수익 ÷ 그 수익을 낸 투자금액 × 100.
            # 금액을 못 낸 건은 분자·분모에서 함께 빠져 비율이 흔들리지 않는다.
            # 전 건이 산출되면 분모는 '대상 투자금액'과 같다.
            "return_rate": round(profit / base * 100, 2) if base else None,
        }

    return render(request, "core/calendar.html", {
        "year": year, "month": month, "weeks": weeks,
        "prev_y": prev_y, "prev_m": prev_m, "next_y": next_y, "next_m": next_m,
        "event_count": sum(len(v) for v in events.values()),
        # 이 달에 상환이 확정된 건수(상환 회차만 센다 — 한 투자는 한 번 상환된다)
        "done_count": sum(1 for v in events.values() for e in v if e["done_label"]),
        "summary": summary,
        "active_nav": "calendar",
    })


# ── 엑셀 업로드 ──────────────────────────────────
@family_required
def upload_excel(request):
    """ELS_Curator가 만든 청약중인상품_*.xlsx를 웹에서 업로드해 임포트."""
    import io
    import os as _os

    from django.conf import settings as _settings
    from django.core.management import call_command

    result = None
    if request.method == "POST":
        f = request.FILES.get("excel")
        if not f:
            messages.error(request, "파일을 선택해주세요.")
        elif not f.name.lower().endswith((".xlsx", ".xlsm")):
            messages.error(request, "엑셀 파일(.xlsx)만 업로드할 수 있습니다.")
        elif f.size > 20 * 1024 * 1024:
            messages.error(request, "20MB 이하 파일만 가능합니다.")
        else:
            _os.makedirs(_settings.UPLOAD_DIR, exist_ok=True)
            save_path = _os.path.join(_settings.UPLOAD_DIR, f.name)
            with open(save_path, "wb") as dest:
                for chunk in f.chunks():
                    dest.write(chunk)

            out = io.StringIO()
            try:
                call_command("import_els", file=save_path, stdout=out)
                result = out.getvalue().strip() or "처리 완료"
                # 새 상품이 들어왔으니 레이더 신호 주차 캐시 무효화
                # (안 하면 이번 주 배지·TOP5가 다음날까지 옛 데이터 기준)
                from core.models import _RADAR_POOL_CACHE
                _RADAR_POOL_CACHE.clear()
                messages.success(request, f"'{f.name}' 임포트 완료")
            except Exception as e:  # noqa: BLE001
                messages.error(request, f"임포트 오류: {e}")

    recent = ImportLog.objects.order_by("-imported_at")[:10]
    return render(request, "core/upload.html", {
        "result": result,
        "recent": recent,
        "active_nav": "upload",
    })


# ── 회원 관리 (운영자 전용) ─────────────────────────
@admin_required
def member_admin(request):
    from django.contrib.auth import get_user_model
    User = get_user_model()

    if request.method == "POST":
        target = get_object_or_404(User, pk=request.POST.get("id"))
        action = request.POST.get("action")
        if target.is_superuser:
            messages.error(request, "운영자 계정은 변경할 수 없습니다.")
        elif action == "toggle_active":
            target.is_active = not target.is_active
            target.save()
            messages.success(request, f"{target.username} 계정을 {'활성화' if target.is_active else '비활성화'}했습니다.")
        elif action == "toggle_staff":
            target.is_staff = not target.is_staff
            target.save()
            messages.success(request, f"{target.username} 계정을 {'가족(staff)으로 지정' if target.is_staff else '일반회원으로 변경'}했습니다.")
        return redirect("member_admin")

    # 회원 수가 적어 파이썬 집계 (다중 조인 annotate의 Sum 부풀림 회피)
    members = list(User.objects.order_by("-date_joined"))
    for m in members:
        held = m.investments.filter(status="보유중")
        m.inv_count = held.count()
        m.inv_total = sum(i.amount for i in held)
        m.watch_count = m.watch_items.count()
        m.preset_count = m.presets.count()
    return render(request, "core/members.html", {
        "members": members,
        "active_nav": "members",
    })


# ── PWA (홈화면 앱 설치) ─────────────────────────
def pwa_manifest(request):
    from django.http import JsonResponse
    return JsonResponse({
        "name": "ELS 레이더",
        "short_name": "ELS 레이더",
        "description": "매주 쏟아지는 ELS, 레이더가 대신 찾아드립니다",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#F2F4F6",
        "theme_color": "#3182F6",
        "icons": [
            {"src": "/icons/icon-192.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/icons/icon-512.png", "sizes": "512x512", "type": "image/png",
             "purpose": "any maskable"},
        ],
    })


def pwa_icon(request, size):
    from pathlib import Path

    from django.conf import settings as _s
    from django.http import FileResponse, Http404
    if size not in ("180", "192", "512"):
        raise Http404
    path = Path(_s.BASE_DIR) / "core" / "assets" / f"icon-{size}.png"
    if not path.exists():
        raise Http404
    resp = FileResponse(open(path, "rb"), content_type="image/png")
    resp["Cache-Control"] = "public, max-age=86400"
    return resp


def threads_card(request, day):
    """스레드 게시용 데이터 카드 이미지.

    Meta 서버가 게시 시점에 이 URL을 직접 가져간다. 인증을 걸면 안 되고,
    캐시를 길게 줘야 재게시 때 원본 서버 부담이 없다.
    """
    from pathlib import Path

    from django.conf import settings as _s
    from django.http import FileResponse, Http404
    path = Path(_s.BASE_DIR) / "core" / "assets" / "threads" / f"card_{day:02d}.png"
    if not path.exists():
        raise Http404
    resp = FileResponse(open(path, "rb"), content_type="image/png")
    resp["Cache-Control"] = "public, max-age=604800"
    return resp


def csrf_failure(request, reason=""):
    """CSRF 검증 실패 커스텀 처리 (settings.CSRF_FAILURE_VIEW).

    로그인 성공이 CSRF 토큰을 회전시킨 직후 같은 폼이 다시 제출되면
    (모바일 연타·브라우저 재전송) 옛 토큰이라 403이 난다. 이 경우
    사용자는 이미 로그인돼 있으므로 에러 대신 홈으로 보낸다.
    비로그인 실패(오래 열린 페이지 등)만 새로고침 안내 페이지를 보여준다.
    """
    if request.user.is_authenticated:
        return redirect(settings.LOGIN_REDIRECT_URL)
    return render(request, "core/csrf_failure.html", status=403)


# ── 웹 푸시 구독 ─────────────────────────────────
def service_worker(request):
    """서비스 워커 스크립트 — 루트(/sw.js)에서 서빙해야 스코프가 사이트 전체가 된다."""
    from pathlib import Path

    from django.http import FileResponse, Http404
    path = Path(settings.BASE_DIR) / "core" / "assets" / "sw.js"
    if not path.exists():
        raise Http404
    resp = FileResponse(open(path, "rb"), content_type="application/javascript")
    resp["Cache-Control"] = "no-cache"  # 워커 수정이 다음 방문에 바로 반영되도록
    return resp


def robots_txt(request):
    """검색엔진 크롤링 규칙 + 사이트맵 위치.

    ⚠ 이 응답이 실제로 나가는지는 배포 후 확인해야 한다. 현재 elsrader.site는
    Cloudflare가 관리형 robots.txt를 내주고 있고(그 응답에는 Sitemap 지시어가
    없다), Cloudflare가 우리 응답을 덮으면 이 뷰는 도달하지 않는다.
    확인: curl -s https://elsrader.site/robots.txt 에 아래 Sitemap 줄이 보이는지.
    안 보이면 Cloudflare 대시보드에서 관리형 robots.txt를 끄거나, Search Console에
    사이트맵 주소를 직접 제출해야 한다.

    로그인·개인 데이터 화면은 크롤러가 들어가 봐야 로그인 페이지만 보므로 막는다.
    막아도 색인에서 사라지는 것은 아니지만(그건 noindex의 일), 크롤링 예산이
    유입을 만드는 페이지로 몰린다.
    """
    from django.http import HttpResponse
    from django.urls import reverse

    sitemap_url = request.build_absolute_uri(reverse("django.contrib.sitemaps.views.sitemap"))
    lines = [
        "User-agent: *",
        "Allow: /",
        "",
        "# 로그인이 필요하거나 개인 데이터가 있는 화면",
        "Disallow: /accounts/",
        "Disallow: /portfolio/",
        "Disallow: /watchlist/",
        "Disallow: /calendar/",
        "Disallow: /presets/",
        "Disallow: /ask/",
        "Disallow: /admin/",
        "",
        "# 운영 전용",
        "Disallow: /manage/",
        "Disallow: /upload/",
        "Disallow: /stats/",
        "",
        "# 내부 검색 결과 — 색인 가치가 없고 무한히 늘어난다",
        "Disallow: /search/",
        "",
        f"Sitemap: {sitemap_url}",
        "",
    ]
    return HttpResponse("\n".join(lines), content_type="text/plain; charset=utf-8")


# 네이버 서치어드바이저 사이트 소유확인 파일 (2026-08-24).
# 파일을 정적으로 두지 않고 뷰로 내는 이유: 이 프로젝트는 정적 파일 서빙 경로를
# 따로 두지 않아 루트에 파일 하나 올리는 게 오히려 번거롭다.
# 값은 비밀이 아니다 — 네이버가 이 주소를 그대로 읽어 대조하는 용도다.
NAVER_VERIFY_FILE = "naver4dd9d4b93cb376cada118943e077602d.html"


def naver_verify(request):
    """/naver<해시>.html — 네이버가 이 파일을 읽어 소유를 확인한다."""
    return HttpResponse(f"naver-site-verification: {NAVER_VERIFY_FILE}",
                        content_type="text/html; charset=utf-8")


def _push_login_required(view):
    """푸시 엔드포인트 전용 로그인 검사.

    이 세 뷰는 fetch()만 호출한다. login_required는 미로그인 시 로그인 페이지로
    302 리다이렉트하는데, fetch는 리다이렉트를 따라가 200(로그인 HTML)을 받는다.
    그러면 클라이언트의 res.ok가 true가 되어 저장되지도 않은 구독을 '성공'으로
    표시한다 — 리다이렉트 대신 403 JSON을 돌려줘야 클라이언트가 실패를 안다.
    """
    from functools import wraps

    from django.http import JsonResponse

    @wraps(view)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return JsonResponse({"ok": False, "error": "auth"}, status=403)
        return view(request, *args, **kwargs)

    return _wrapped


@_push_login_required
@require_POST
def push_subscribe(request):
    """브라우저가 발급받은 푸시 구독 저장. 같은 endpoint 재등록 시 소유자·키 갱신."""
    import json

    from django.http import JsonResponse

    from .models import PushSubscription
    try:
        data = json.loads(request.body)
        endpoint = data["endpoint"]
        p256dh = data["keys"]["p256dh"]
        auth = data["keys"]["auth"]
    except (ValueError, KeyError, TypeError):
        return JsonResponse({"ok": False}, status=400)
    if not endpoint.startswith("https://"):
        return JsonResponse({"ok": False}, status=400)
    PushSubscription.objects.update_or_create(
        endpoint=endpoint,
        defaults={"user": request.user, "p256dh": p256dh, "auth": auth,
                  "user_agent": request.META.get("HTTP_USER_AGENT", "")[:200]},
    )
    return JsonResponse({"ok": True})


@_push_login_required
@require_POST
def push_unsubscribe(request):
    """알림 끄기 — endpoint 소유(브라우저)가 곧 증명이라 endpoint 기준으로 삭제."""
    import json

    from django.http import JsonResponse

    from .models import PushSubscription
    try:
        endpoint = json.loads(request.body)["endpoint"]
    except (ValueError, KeyError, TypeError):
        return JsonResponse({"ok": False}, status=400)
    PushSubscription.objects.filter(endpoint=endpoint).delete()
    return JsonResponse({"ok": True})


@_push_login_required
@require_POST
def push_test(request):
    """구독 직후 확인용 테스트 알림 1건."""
    from django.http import JsonResponse

    from . import push
    n = push.send_to_user(
        request.user, "알림이 켜졌어요",
        "조기상환 평가 전날, 낙인 근접, 관심상품 청약마감 전날 알려드립니다.",
        url="/portfolio/",
    )
    return JsonResponse({"ok": n > 0, "sent": n})


def og_image(request):
    """링크 공유 미리보기 이미지 (1200x630, base.html og:image가 참조)."""
    from pathlib import Path

    from django.conf import settings as _s
    from django.http import FileResponse, Http404
    path = Path(_s.BASE_DIR) / "core" / "assets" / "og.png"
    if not path.exists():
        raise Http404
    resp = FileResponse(open(path, "rb"), content_type="image/png")
    resp["Cache-Control"] = "public, max-age=86400"
    return resp


# ── 상품 검색 (공개) ─────────────────────────────
# 자연어 조건검색은 /ask/로 옮겼다. 여기는 키워드 검색만 남긴다 —
# 같은 일을 하는 입구가 둘이면 어느 쪽이 최신인지 아무도 모르게 된다.
def product_search(request):
    from django.db.models import Q

    q = request.GET.get("q", "").strip()
    results = []
    if q:
        results = list(
            Product.objects.listed().filter(
                Q(product_no__icontains=q) | Q(issuer__icontains=q)
                | Q(assets_raw__icontains=q) | Q(name__icontains=q)
            ).order_by("-sub_end")[:200]
        )

    invested_ids = set()
    if request.user.is_authenticated:
        invested_ids = set(Investment.objects.filter(
            user=request.user, status="보유중").values_list("product_id", flat=True))
    return render(request, "core/search.html", {
        "q": q, "results": results, "invested_ids": invested_ids,
        "active_nav": "search",
    })


# ── AI 분석 질문 (/ask/) ─────────────────────────
# 예시 질문 — 네 갈래(시세·발행통계·조건검색·포트폴리오)를 한 화면에서 보여준다
ASK_PRESETS = [
    {"label": "기초자산 시세", "icon": "fa-chart-area", "private": False, "items": [
        "Tesla와 삼성전자 10년 최대낙폭 비교",
        "2020년 3월에 가장 많이 빠진 기초자산은",
        "코스피200이 지금 5년 고점 대비 몇 %야"]},
    {"label": "ELS 발행·상환 통계", "icon": "fa-chart-simple", "private": False, "items": [
        "연도별 ELS 실현수익률과 손실 건수",
        "올해 가장 많이 쓰인 기초자산",
        "낙인 구간별 발행 비중 5년 추이"]},
    {"label": "조건 검색", "icon": "fa-filter", "private": False, "items": [
        "낙인 40 이하 지수형 중 수익률 높은 3개",
        "이번 주 마감인 노낙인 상품",
        "Micron 들어간 청약중 상품 전부"]},
    {"label": "내 포트폴리오", "icon": "fa-briefcase", "private": True, "items": [
        "내 보유 중 낙인까지 여유가 가장 적은 건",
        "기초자산별 비중이랑 집중도 정리해줘",
        "다음 달 평가일 오는 상품 몇 건이야"]},
]

ASK_DISCLAIMERS = {
    "market": "과거 데이터를 집계한 결과이며 투자 권유가 아닙니다. 과거 성과는 미래 수익을 "
              "보장하지 않고, 원금 손실이 발생할 수 있습니다.",
    "portfolio": "등록하신 보유 기록을 집계한 결과이며 투자 권유가 아닙니다. 스트레스 결과는 "
                 "가정에 따른 계산일 뿐 예측이 아니고, 원금 손실이 발생할 수 있습니다.",
}


def _ask_limit(request_user):
    """(하루 한도, 면제 사유). 한도가 None이면 무제한.

    ⚠ is_staff 는 '가족'(엑셀 업로드 권한을 가진 일반 회원)이고
      is_superuser 가 운영자다. 무제한은 운영자에게만 연다 — 가족은 서비스
      사용자지 운영자가 아니고, 넓게 열면 비용이 예측 밖으로 나간다.
    """
    limits = getattr(settings, "ASK_DAILY_LIMITS", {"default": 3})
    if request_user.is_superuser:
        return limits.get("superuser", None), "superuser"
    if request_user.is_staff:
        return limits.get("staff", limits.get("default", 3)), "staff"
    return limits.get("default", 3), None


def _ask_quota(user):
    from .models import AskLog

    limit, why = _ask_limit(user)
    today = timezone.localdate()
    used = AskLog.objects.filter(user=user, asked_on=today, billed=True).count()
    unlimited = limit is None
    return {
        "used": used, "limit": limit, "unlimited": unlimited,
        "exempt_reason": why if unlimited else None,
        "remaining": None if unlimited else max(0, limit - used),
        "exhausted": (not unlimited) and used >= limit,
        "percent": 100 if unlimited else min(100, round(used / limit * 100)) if limit else 0,
        "reset_text": "내일 오전 0시에 다시 채워집니다.",
    }


def _ask_context(request, result=None, quota=None, from_cache=False, log_id=None):
    from .models import AskLog

    today = timezone.localdate()
    history = [
        {"id": r.id, "question": r.question, "at": timezone.localtime(r.created_at),
         "status": r.status, "tools": r.tools_used}
        for r in AskLog.objects.filter(user=request.user, asked_on=today)
        .exclude(status=AskLog.STATUS_BLOCKED).order_by("-created_at")[:10]
    ]
    ctx = {
        "active_nav": "ask",
        "presets": ASK_PRESETS,
        "coverage": ask_tools.coverage(),
        "quota": quota or _ask_quota(request.user),
        "history": history,
        "from_cache": from_cache,
        "log_id": log_id,
        "ai_enabled": bool(settings.ANTHROPIC_API_KEY),
        # 답변 영역 (없으면 템플릿이 입력 폼만 그린다)
        "question": None, "answer": None, "answer_html": None,
        "blocks": [], "basis": None,
        "followups": [], "error": None, "status": None, "tools": [],
        "elapsed_ms": None, "elapsed_s": None, "answered_at": None,
        "disclaimer": None, "is_private": False,
    }
    if result:
        ms = result.get("elapsed_ms")
        ctx.update({
            "question": result.get("question"),
            "answer": result.get("answer") or None,
            # 도구 표시값만 감싼 강조 HTML (ask_agent.highlight) — |safe 로 쓴다
            "answer_html": result.get("answer_html") or None,
            "blocks": result.get("blocks") or [],
            "basis": result.get("basis"),
            "followups": result.get("followups") or [],
            "error": result.get("error"),
            "status": result.get("status"),
            "tools": result.get("tools") or [],
            "elapsed_ms": ms,
            "elapsed_s": round(ms / 1000.0, 1) if ms else None,
            "answered_at": result.get("answered_at"),
            "is_private": "portfolio_facts" in (result.get("tools") or []),
        })
        ctx["disclaimer"] = ASK_DISCLAIMERS["portfolio" if ctx["is_private"] else "market"]
    return ctx


@login_required
def ask(request):
    """AI 분석 질문 — 로그인 필수, 하루 한도, 당일 같은 질문은 저장된 답을 재사용."""
    from . import ask_agent
    from .models import AskLog

    today = timezone.localdate()

    # 저장된 답 다시 열기 (본인 것만, 차감 없음)
    log_id = request.GET.get("log")
    if request.method == "GET" and log_id:
        log = AskLog.objects.filter(pk=log_id, user=request.user).first()
        if log and log.payload:
            return render(request, "core/ask.html",
                          _ask_context(request, log.payload, from_cache=True, log_id=log.id))
        return render(request, "core/ask.html", _ask_context(request))

    if request.method != "POST":
        return render(request, "core/ask.html", _ask_context(request))

    q = ask_agent.normalize_question(request.POST.get("q", ""))
    if not q:
        return redirect("ask")

    # 질문키에는 시세 최신일이 섞인다 — 09:30 배치 전 답이 배치 뒤에 재사용되면
    # 화면과 답이 어긋난다 (ask_agent.question_key 참조)
    key = ask_agent.question_key(q)

    def _blocked(code, message, quota=None):
        AskLog.objects.create(
            user=request.user, asked_on=today, question=q, question_key=key,
            status=AskLog.STATUS_BLOCKED, refuse_code=code, billed=False,
            payload={}, answer="")
        ctx = _ask_context(request, quota=quota or _ask_quota(request.user))
        ctx["error"] = {"code": code, "message": message, "detail": message}
        ctx["question"] = q
        return render(request, "core/ask.html", ctx)

    # ① 당일 캐시 — 같은 질문은 모델을 다시 부르지 않는다 → 차감도 없다.
    #    ⚠ 포트폴리오 질문은 제외한다. 보유·시세가 하루 안에도 움직여서
    #      아침에 만든 답을 오후에 다시 보여주면 그 순간 틀린 화면이 된다.
    hit = (AskLog.objects.filter(user=request.user, asked_on=today, question_key=key,
                                 status__in=[AskLog.STATUS_OK, AskLog.STATUS_GUARDED])
           .exclude(payload={})
           .exclude(tools_used__icontains="portfolio_facts")
           .order_by("-created_at").first())
    if hit:
        return render(request, "core/ask.html",
                      _ask_context(request, hit.payload, from_cache=True, log_id=hit.id))

    # ② 킬스위치 — 사고 시 배포 없이 끈다
    if not getattr(settings, "ASK_ENABLED", True):
        return _blocked("DISABLED", "AI 분석 질문을 잠시 중단했습니다. 곧 다시 열겠습니다.")

    # ③ 서비스 전체 하루 상한 — 계정을 여러 개 만들어 1인 한도를 우회하는 경우 대비
    cap = getattr(settings, "ASK_GLOBAL_DAILY_CAP", 0)
    if cap and AskLog.objects.filter(asked_on=today, billed=True).count() >= cap:
        return _blocked("GLOBAL_CAP",
                        "오늘 서비스 전체 질문 한도에 도달했습니다. 내일 다시 이용해 주세요.")

    # ④ 1인 한도 — 면제 계정은 체크를 건너뛴다 (기록은 남긴다)
    quota = _ask_quota(request.user)
    if quota["exhausted"]:
        return _blocked("QUOTA_EXCEEDED",
                        f"오늘 질문 {quota['limit']}회를 모두 사용했습니다.", quota)

    if not settings.ANTHROPIC_API_KEY:
        ctx = _ask_context(request)
        ctx["question"] = q
        ctx["error"] = {"code": "NOT_CONFIGURED", "message": "AI 분석 질문이 아직 설정되지 않았습니다."}
        return render(request, "core/ask.html", ctx)

    # ③ 실행
    result = ask_agent.run(request.user, q)
    result["question"] = q
    result["answered_at"] = timezone.localtime().isoformat(timespec="minutes")

    # ④ 차감 규칙
    #    정상 O / 사후검사 실패 O(비용이 실제로 발생) / 캐시 히트 X(위에서 끝남)
    #    사전 거절 X / 오류 X / 무제한 계정 X — 다만 전부 기록은 남긴다.
    billable = result["status"] in (AskLog.STATUS_OK, AskLog.STATUS_GUARDED)
    usage = result.get("usage") or {}
    tot = {k: sum(u.get(k, 0) for u in usage.values())
           for k in ("input_tokens", "output_tokens",
                     "cache_creation_input_tokens", "cache_read_input_tokens")}
    payload = {k: result.get(k) for k in
               ("question", "answer", "answer_html", "blocks", "basis", "followups",
                "tools", "error", "status", "elapsed_ms", "answered_at")}
    log = AskLog.objects.create(
        user=request.user, asked_on=today, question=q, question_key=key,
        status=result["status"], refuse_code=(result.get("error") or {}).get("code", "") or "",
        guard_flags=result.get("guard_flags") or [],
        answer=result.get("answer") or "", tools_used=result.get("tools") or [],
        payload=payload if billable else {},
        input_tokens=tot["input_tokens"], output_tokens=tot["output_tokens"],
        cache_write_tokens=tot["cache_creation_input_tokens"],
        cache_read_tokens=tot["cache_read_input_tokens"],
        cost_usd=result.get("cost_usd") or 0.0, usage=usage,
        elapsed_ms=result.get("elapsed_ms") or 0,
        billed=billable and not quota["unlimited"],
    )
    return render(request, "core/ask.html",
                  _ask_context(request, result, quota=_ask_quota(request.user), log_id=log.id))


# ── 홈 (/) — 로그인 여부 무관 랜딩 ──
def home(request):
    return about(request)


# ── 소개 랜딩 (공개) ─────────────────────────────
_ABOUT_CACHE = {"day": None, "ctx": None}


def _about_accuracy():
    """레이더 신호 성과검증 집계 — 배지 상품 vs 배지 없는 상품의 평가일 통과율."""
    from django.db.models import Count, Max, Min, Q

    from .models import PRINCIPAL_PROTECTED_TYPES, RadarVerdict

    # 트렌드 화면과 같은 모수를 쓴다 — 원금지급형은 대조군에서도 뺀다
    qs = (RadarVerdict.objects.filter(met__isnull=False)
          .exclude(product__product_type__in=PRINCIPAL_PROTECTED_TYPES))
    rows = qs.values("tier").annotate(n=Count("id"), ok=Count("id", filter=Q(met=True)))
    badge_n = badge_ok = ctrl_n = ctrl_ok = 0
    for r in rows:
        if r["tier"] in ("타겟 신호", "아주 강한 신호", "강한 신호"):
            badge_n += r["n"]
            badge_ok += r["ok"]
        else:
            ctrl_n += r["n"]
            ctrl_ok += r["ok"]
    if not badge_n or not ctrl_n:
        return None
    span = qs.aggregate(a=Min("week_monday"), b=Max("week_monday"))
    badge_pct = badge_ok / badge_n * 100
    ctrl_pct = ctrl_ok / ctrl_n * 100
    return {
        "badge_n": badge_n, "badge_ok": badge_ok, "badge_pct": round(badge_pct, 1),
        "ctrl_n": ctrl_n, "ctrl_ok": ctrl_ok, "ctrl_pct": round(ctrl_pct, 1),
        "delta": round(badge_pct - ctrl_pct, 1),
        "total_n": badge_n + ctrl_n, "start": span["a"], "end": span["b"],
        # 막대 높이(%) — 60% 기준선 축으로 차이를 시각적으로 벌린다 (값은 라벨에 명시)
        "badge_h": round(max(8.0, (badge_pct - 60) / 40 * 90), 1),
        "ctrl_h": round(max(8.0, (ctrl_pct - 60) / 40 * 90), 1),
    }


def _about_ki_series():
    """SEIBro 표본의 연도별 평균 낙인 배리어 → 꺾은선 좌표(viewBox 620x190)."""
    from django.db.models import Avg, Count
    from django.db.models.functions import ExtractYear

    from .models import HistoricalIssue

    rows = list(HistoricalIssue.objects.filter(ki__isnull=False)
                .annotate(y=ExtractYear("issue_date")).values("y")
                .annotate(n=Count("id"), avg=Avg("ki")).order_by("y"))
    rows = [r for r in rows if r["y"] and r["n"] >= 30]   # 표본 30건 미만 연도는 제외
    if len(rows) < 5:
        return None

    x0, x1, y_top, y_bot = 46.0, 596.0, 22.0, 150.0      # 그리기 영역
    lo, hi = 40.0, 62.0                                   # y축 범위(%)
    step = (x1 - x0) / (len(rows) - 1)

    def _y(v):
        return round(y_bot - (v - lo) / (hi - lo) * (y_bot - y_top), 1)

    pts = []
    for i, r in enumerate(rows):
        pts.append({"x": round(x0 + step * i, 1), "y": _y(r["avg"]),
                    "year": r["y"], "avg": round(r["avg"], 1), "n": r["n"]})
    return {
        "points": " ".join(f"{p['x']},{p['y']}" for p in pts),
        "first": pts[0], "last": pts[-1], "marks": [pts[0], pts[len(pts) // 2], pts[-1]],
        "sample_n": sum(r["n"] for r in rows),
        "grid": [{"v": v, "y": _y(v)} for v in (60, 55, 50, 45)],
        "drop": round(max(p["avg"] for p in pts) - pts[-1]["avg"], 1),
    }


def _about_live():
    """운영자 실계좌 상환 실적 집계 — 금액은 노출하지 않고 비율·건수만."""
    done = [i for i in Investment.objects.filter(
        status__in=["조기상환", "만기상환", "낙인후상환"],
        redeemed_amount__isnull=False, redeemed_at__isnull=False)
        if i.amount and (i.redeemed_at - i.invested_at).days > 0]
    if not done:
        return None
    anns, months, wins = [], [], 0
    for i in done:
        days = (i.redeemed_at - i.invested_at).days
        r = (i.redeemed_amount - i.amount) / i.amount
        anns.append(r * 365 / days)
        months.append(days / 30.4)
        if r > 0:
            wins += 1
    n = len(done)
    return {"n": n, "win": wins,
            "ann": round(sum(anns) / n * 100, 1),
            "avg_m": round(sum(months) / n, 1)}


def about(request):
    """소개 랜딩 — 실측 숫자는 하루 1회만 계산해 캐시(마케팅 유입 대비)."""
    today = date.today()
    if _ABOUT_CACHE["day"] == today and _ABOUT_CACHE["ctx"]:
        ctx = _ABOUT_CACHE["ctx"]
    else:
        from .models import radar_top5
        top5 = radar_top5()
        if not top5:                       # 마감이 지나 이번주가 비면 지난주 것으로
            last_mon = today - timedelta(days=today.weekday() + 7)
            top5 = radar_top5(last_mon, last_mon + timedelta(days=6))
        # stat_*·live는 랜딩 개편(2026-07-31)으로 화면에서 빠짐 — _about_live는 실전 인증
        # 섹션 재삽입 때 다시 쓸 예정이라 헬퍼는 남겨둔다.
        ctx = {
            "accuracy": _about_accuracy(),
            "ki": _about_ki_series(),
            "top5": top5,
        }
        _ABOUT_CACHE.update(day=today, ctx=ctx)
    # /about/ 는 home 과 완전히 같은 HTML을 낸다(home 이 이 함수를 그대로 호출).
    # 둘을 각각 색인시키면 같은 페이지가 둘로 갈라져 어느 쪽도 순위를 못 받는다.
    # 대표 주소를 / 하나로 모으고, 사이트맵에도 / 만 넣는다(core/sitemaps.py).
    return render(request, "core/about.html", {
        "active_nav": "about",
        "canonical_url": request.build_absolute_uri("/"),
        **ctx,
    })


# ── 리포트 (공개) ─────────────────────────────────
def report_els_10year(request):
    """ELS 10년 성적표 — 69,903건 전수 판정 (콘텐츠 1편).

    본문은 템플릿에 원고 그대로 박아 둔다. DB 집계로 다시 뽑지 않는다 —
    원고의 수치는 2026-08-13 기준으로 검증·컨펌을 마친 확정값이고, 매 요청마다
    재계산하면 검증받지 않은 숫자가 조용히 화면에 올라간다. 갱신할 때는
    원고와 core/sitemaps.py 의 REPORT_10YEAR_UPDATED 를 함께 고친다.
    """
    from .sitemaps import REPORT_10YEAR_UPDATED

    return render(request, "core/report_els_10year.html", {
        "updated": REPORT_10YEAR_UPDATED,
        "meta_desc": ("2016~2025년 공모 ELS 69,903건을 전수 판정했더니 원금손실은 3,220건, "
                      "정상상환율 95.4%였습니다. 연도별 손실률과 판정 기준·한계까지 공개합니다."),
    })


# ── 법적 페이지 (공개) ────────────────────────────
def legal_terms(request):
    """이용약관."""
    return render(request, "core/legal_terms.html", {
        "meta_desc": "ELS 레이더 서비스의 이용 조건과 운영자·이용자의 권리·의무를 규정한 이용약관입니다.",
    })


def legal_privacy(request):
    """개인정보처리방침."""
    return render(request, "core/legal_privacy.html", {
        "meta_desc": ("ELS 레이더가 수집하는 개인정보 항목과 이용 목적, 보관 기간, "
                      "이용자의 권리를 안내합니다."),
    })


def legal_disclaimer(request):
    """투자 유의사항(면책)."""
    return render(request, "core/legal_disclaimer.html", {
        "meta_desc": ("ELS 레이더가 제공하는 손실확률·신호 등급의 산출 근거와 한계, "
                      "그리고 ELS 투자 전 반드시 알아야 할 원금손실 위험을 정리했습니다."),
    })


# ── 베타 피드백 ───────────────────────────────────
def should_show_feedback_banner(user) -> bool:
    """이 사용자에게 베타 피드백 배너를 띄울지.

    안 띄우는 경우
      ① 비로그인 — 배너의 목적이 '쓰고 있는 사람의 불편'을 받는 것이라
         가입도 안 한 사람에게 물을 게 없다.
      ② 이미 의견을 보냈다 — 목적을 달성했다.
      ③ 직접 닫았다 — 아래 재노출 정책 참고.

    ▣ 재노출 정책: 닫으면 영구 숨김 (2026-08-13 결정)
      사용자가 극소수인 지금은 '한 번 더 띄워 얻는 의견 1건'보다 '또 뜨는 배너에
      질려 떠나는 사람 1명'의 손실이 크다. 닫은 뒤에도 푸터의 '의견 보내기'가
      상시 남으므로 경로 자체가 사라지지는 않는다.
      30일 재노출로 바꾸려면 ③ 조건만 dismissed_at 기준 비교로 고치면 된다
      (FeedbackBannerDismissal이 시각을 이미 저장하고 있다).
    """
    if not user.is_authenticated:
        return False
    if Feedback.objects.filter(user=user).exists():
        return False
    return not FeedbackBannerDismissal.objects.filter(user=user).exists()


def _feedback_notify(fb):
    """피드백 접수를 텔레그램으로 알린다. 실패해도 저장은 유지된다.

    ⚠ 연락처 원문을 싣지 않는다. 텔레그램은 chat_id 하나짜리 공용 채널이라
      운영자 외의 사람도 읽는다 — 사용자 전화번호·카톡 아이디가 그 방에
      영구히 남는 건 받은 목적(기프티콘 발송)에 비해 과하다. 남겼는지 여부만
      알리고 실제 값은 관리자 화면에서 본다.
    """
    if not settings.TELEGRAM_FEEDBACK_ALERT_ENABLED:
        return False
    lines = [
        f"[베타 피드백] {fb.user.username}",
        f"인터뷰(30분 통화): {'가능' if fb.interview_ok else '응답 없음'}",
        f"연락처: {'남김 — 관리자 화면에서 확인' if fb.contact else '없음(가입 이메일로 회신)'}",
        "",
        fb.body,
        "",
        f"{settings.SITE_URL.rstrip('/')}/admin/core/feedback/{fb.id}/change/",
    ]
    return telegram.send_message("\n".join(lines))


@login_required
def feedback(request):
    """의견 보내기 — 자유 텍스트 1칸 + 연락처(선택) + 인터뷰 의향.

    항목을 더 늘리지 않는다. 베타에서 필요한 건 '무엇이 불편했나' 한 문장이고,
    칸이 늘수록 제출률이 떨어진다. 별점·분류 같은 건 표본이 쌓인 뒤에 붙인다.
    """
    from django.urls import reverse

    if request.method == "POST":
        body = (request.POST.get("body") or "").strip()
        contact = (request.POST.get("contact") or "").strip()[:120]
        interview_ok = bool(request.POST.get("interview_ok"))
        if not body:
            return render(request, "core/feedback.html", {
                "error": "의견을 한 줄이라도 적어주세요.",
                "body": body, "contact": contact, "interview_ok": interview_ok,
                "body_max": Feedback.BODY_MAX, "active_nav": "feedback",
            })
        fb = Feedback.objects.create(
            user=request.user, body=body[:Feedback.BODY_MAX],
            contact=contact, interview_ok=interview_ok,
        )
        try:
            _feedback_notify(fb)
        except Exception as e:
            # 알림 실패로 사용자 제출을 되돌리지 않는다 — 기록은 이미 남았고
            # 관리자 화면에서 볼 수 있다.
            logger.error("피드백 텔레그램 알림 실패: %s", e)
        # 완료 화면으로 리다이렉트(PRG). 토스트 한 줄로 끝내지 않는 이유는,
        # 인터뷰를 신청한 사람에게 '언제 어디로 연락이 오는지'를 말해줘야 하기
        # 때문이다 — 지금 이 폼이 유일한 인터뷰 모집 경로다.
        # 리다이렉트로 갈라두면 새로고침해도 같은 의견이 두 번 저장되지 않는다.
        return redirect(f"{reverse('feedback')}?sent=1")

    # ?sent=1은 주소만 알면 누구나 붙일 수 있으므로, 실제로 보낸 기록이
    # 있을 때만 완료 화면을 보여준다. 없으면 그냥 폼이다.
    sent = None
    if request.GET.get("sent"):
        sent = Feedback.objects.filter(user=request.user).order_by("-created_at").first()
    return render(request, "core/feedback.html",
                  {"sent": sent, "body_max": Feedback.BODY_MAX, "active_nav": "feedback"})


@login_required
@require_POST
def feedback_dismiss(request):
    """배너 닫기 — 계정에 기록해 다시 뜨지 않게 한다.

    fetch로 오면 JSON만 주고 배너를 그 자리에서 지운다(화면 새로고침 없음).
    JS가 없거나 실패하면 평범한 폼 제출로 들어와 원래 보던 화면으로 돌아간다
    — 별표 토글(base.html)과 같은 방식이다.
    """
    FeedbackBannerDismissal.objects.get_or_create(user=request.user)
    if _pf_is_fetch(request):
        return JsonResponse({"dismissed": True})
    # next는 사용자가 보내는 값이라 그대로 믿지 않는다 — 같은 사이트 주소일
    # 때만 돌아간다(외부로 튕기는 열린 리다이렉트 방지).
    nxt = request.POST.get("next") or ""
    if nxt and url_has_allowed_host_and_scheme(
            nxt, allowed_hosts={request.get_host()}, require_https=request.is_secure()):
        return redirect(nxt)
    return redirect("weekly")


# ── 회원 탈퇴 ────────────────────────────────────
@login_required
def account_delete(request):
    """계정 및 관련 데이터 완전 삭제.

    Investment / WatchItem / Preset 은 모두 user FK on_delete=CASCADE 이므로
    user.delete() 한 번으로 함께 삭제된다.
    실수 방지를 위해 비밀번호 + 확인 문구("탈퇴합니다") 두 가지를 요구한다.
    """
    from django.contrib.auth import logout as auth_logout

    user = request.user
    # 운영자 계정은 이 화면으로 삭제 불가 (실수로 지우면 관리 권한 복구 불가)
    if user.is_superuser:
        messages.error(request, "운영자 계정은 이 화면에서 탈퇴할 수 없습니다.")
        return redirect("weekly")

    # 카카오로만 가입한 계정은 비밀번호가 없다(allauth 가 사용 불가로 둔다).
    # 비밀번호 확인을 그대로 요구하면 그 사람은 영영 탈퇴할 수 없으므로,
    # 확인 문구 하나만으로 진행한다.
    has_password = user.has_usable_password()

    ctx = {
        "inv_count": Investment.objects.filter(user=user).count(),
        "watch_count": WatchItem.objects.filter(user=user).count(),
        "preset_count": Preset.objects.filter(user=user).count(),
        "has_password": has_password,
    }

    if request.method == "POST":
        password = request.POST.get("password", "")
        confirm = (request.POST.get("confirm") or "").strip()
        if has_password and not user.check_password(password):
            ctx["error"] = "비밀번호가 올바르지 않습니다."
            return render(request, "core/account_delete.html", ctx)
        if confirm != "탈퇴합니다":
            ctx["error"] = "확인 문구가 일치하지 않습니다. '탈퇴합니다' 를 그대로 입력해 주세요."
            return render(request, "core/account_delete.html", ctx)

        auth_logout(request)          # 세션 먼저 정리
        user.delete()                 # 계정 + CASCADE 데이터 삭제
        messages.success(request, "탈퇴가 완료되었습니다")
        return redirect("weekly")

    return render(request, "core/account_delete.html", ctx)


# ── 접속 통계 (운영자 전용) ───────────────────────
@admin_required
def stats(request):
    """최근 30일 방문·유입·가입 집계."""
    from django.contrib.auth import get_user_model
    from django.db.models import Count

    from .models import PageView

    today = date.today()
    start = today - timedelta(days=29)
    qs = PageView.objects.filter(date__gte=start)

    daily = {r["date"]: r for r in qs.values("date").annotate(
        views=Count("id"), uniq=Count("visitor", distinct=True))}
    srows = (get_user_model().objects.filter(date_joined__date__gte=start)
             .values("date_joined__date").annotate(n=Count("id")))
    signup_map = {r["date_joined__date"]: r["n"] for r in srows}

    days, max_uniq = [], 1
    for i in range(30):
        d = start + timedelta(days=i)
        r = daily.get(d, {})
        u = r.get("uniq", 0)
        max_uniq = max(max_uniq, u)
        days.append({"d": d, "views": r.get("views", 0), "uniq": u,
                     "signups": signup_map.get(d, 0)})
    for row in days:
        row["h"] = round(row["uniq"] / max_uniq * 100, 1)

    ctx = {
        "days": days, "start": start, "today": today,
        "top_paths": qs.values("path").annotate(n=Count("id")).order_by("-n")[:12],
        "top_refs": (qs.exclude(ref="").values("ref")
                     .annotate(n=Count("visitor", distinct=True)).order_by("-n")[:10]),
        "totals": {
            "views": qs.count(),
            "uniq": qs.values("visitor").distinct().count(),
            "signups": sum(r["signups"] for r in days),
            "today_views": days[-1]["views"], "today_uniq": days[-1]["uniq"],
        },
    }
    return render(request, "core/stats.html", ctx)
