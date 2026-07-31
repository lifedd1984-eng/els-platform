"""연도별 종목형/지수형 100건씩 무작위 샘플 → 엑셀.

재현 가능하도록 난수 시드를 고정한다(seed=20260729).
요약 시트의 수치는 전부 수식으로 넣어, 데이터 시트를 고쳐도 따라 움직이게 한다.
"""
import os
import random

import django
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "els_platform.settings")
django.setup()

from core.models import HistoricalIssue  # noqa: E402

YEARS = [2025, 2024, 2023]
PER_TYPE = 100
SEED = 20260729
OUT = "/home/ubuntu/els/logs/els_sample.xlsx"

HEADERS = [
    ("발행일", 12), ("발행사", 14), ("종목명", 30), ("ISIN", 14),
    ("유형", 8), ("기초자산", 42),
    ("낙인", 7), ("1차배리어", 10), ("막차배리어", 10),
    ("연수익률(%)", 11),
    ("배지등급", 12), ("그룹내순위", 10),
    ("1년내상환성공", 13), ("판정레벨(%)", 11),
    ("시뮬 1년내상환(%)", 15), ("시뮬 손실확률(%)", 15),
]

FONT = "Arial"
hdr_font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1B64DA")
cell_font = Font(name=FONT, size=10)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def asset_names(h):
    out = []
    for a in (h.assets or []):
        n = (a.get("name") if isinstance(a, dict) else str(a)) or ""
        if n:
            out.append(n)
    return " / ".join(out)


def pick(year, is_index):
    qs = HistoricalIssue.objects.filter(
        issue_date__year=year, verdict_met__isnull=False)
    qs = qs.filter(basset_sort="지수") if is_index else qs.exclude(basset_sort="지수")
    ids = list(qs.values_list("id", flat=True))
    rnd = random.Random(f"{SEED}-{year}-{is_index}")
    rnd.shuffle(ids)
    chosen = ids[:PER_TYPE]
    return list(HistoricalIssue.objects.filter(id__in=chosen).order_by("issue_date"))


wb = Workbook()
wb.remove(wb.active)

sheet_names = []
for year in YEARS:
    ws = wb.create_sheet(str(year))
    sheet_names.append(str(year))
    for c, (name, width) in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, name)
        cell.font, cell.fill, cell.border = hdr_font, hdr_fill, border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "A2"

    r = 2
    for is_index in (False, True):
        for h in pick(year, is_index):
            bars = h.stepdown_barriers or []
            vals = [
                h.issue_date, h.issuer, (h.name or "")[:60], h.isin,
                "지수형" if h.basset_sort == "지수" else "종목형",
                asset_names(h),
                h.ki,
                bars[0] if bars and isinstance(bars[0], (int, float)) else None,
                bars[-1] if bars and isinstance(bars[-1], (int, float)) else None,
                h.yield_rate,
                h.radar_tier or "배지없음",
                h.radar_rank,
                "성공" if h.verdict_met else "실패",
                h.verdict_level,
                h.sim_early_1y, h.sim_loss_prob,
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v)
                cell.font, cell.border = cell_font, border
                if c == 1:
                    cell.number_format = "yyyy-mm-dd"
                elif c in (7, 8, 9, 12):
                    cell.number_format = "0"
                elif c in (10, 14, 15, 16):
                    cell.number_format = "0.0"
                if c in (5, 11, 13):
                    cell.alignment = Alignment(horizontal="center")
            r += 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{r - 1}"

# ── 요약 시트 (전부 수식) ──────────────────────────────
ws = wb.create_sheet("요약", 0)
ws.column_dimensions["A"].width = 10
for c in range(2, 8):
    ws.column_dimensions[get_column_letter(c)].width = 15

ws["A1"] = "레이더 검증 샘플 요약"
ws["A1"].font = Font(name=FONT, bold=True, size=13)
ws["A2"] = f"연도별 종목형·지수형 각 {PER_TYPE}건 무작위 추출 (시드 {SEED}, 재현 가능)"
ws["A2"].font = Font(name=FONT, size=9, color="666666")
ws["A3"] = "지표: 1년 이내 조기상환 성공률 = 발행 후 365일 내 평가 회차 중 한 번이라도 배리어 충족"
ws["A3"].font = Font(name=FONT, size=9, color="666666")
ws["A4"] = "2026년은 판정 가능 건수가 최소 표본(300건) 미달이라 제외"
ws["A4"].font = Font(name=FONT, size=9, color="C00000")

heads = ["연도", "유형", "표본", "배지", "배지 성공률", "대조군 성공률", "격차(%p)"]
for c, t in enumerate(heads, 1):
    cell = ws.cell(6, c, t)
    cell.font, cell.fill, cell.border = hdr_font, hdr_fill, border
    cell.alignment = Alignment(horizontal="center")

r = 7
for year in YEARS:
    s = str(year)
    for t in ("종목형", "지수형"):
        ws.cell(r, 1, year).number_format = "0"
        ws.cell(r, 2, t)
        ws.cell(r, 3, f'=COUNTIFS(\'{s}\'!$E:$E,$B{r})')
        ws.cell(r, 4, f'=COUNTIFS(\'{s}\'!$E:$E,$B{r},\'{s}\'!$K:$K,"<>배지없음")')
        ws.cell(r, 5, f'=IFERROR(COUNTIFS(\'{s}\'!$E:$E,$B{r},\'{s}\'!$K:$K,"<>배지없음",'
                      f'\'{s}\'!$M:$M,"성공")/$D{r},"")')
        ws.cell(r, 6, f'=IFERROR(COUNTIFS(\'{s}\'!$E:$E,$B{r},\'{s}\'!$K:$K,"배지없음",'
                      f'\'{s}\'!$M:$M,"성공")/COUNTIFS(\'{s}\'!$E:$E,$B{r},'
                      f'\'{s}\'!$K:$K,"배지없음"),"")')
        ws.cell(r, 7, f'=IFERROR(($E{r}-$F{r})*100,"")')
        for c in range(1, 8):
            cell = ws.cell(r, c)
            cell.font, cell.border = cell_font, border
            if c in (5, 6):
                cell.number_format = "0.0%"
            elif c == 7:
                cell.number_format = "0.0"
            if c in (1, 2):
                cell.alignment = Alignment(horizontal="center")
        r += 1

note = ws.cell(r + 1, 1, "※ 표본은 각 연도 판정확정 건에서 무작위 추출한 것이라, "
                        "전수 검증 수치와 소수점 단위로 다를 수 있습니다.")
note.font = Font(name=FONT, size=9, color="666666")

wb.save(OUT)
print("저장:", OUT)
for s in sheet_names:
    print("  시트", s, "행수", wb[s].max_row - 1)
